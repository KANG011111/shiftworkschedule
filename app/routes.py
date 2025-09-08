from flask import Blueprint, render_template, request, jsonify, redirect, url_for, g, flash, session, Response
from app.models import db, Employee, ShiftType, Schedule, User, ImportLog, GroupMembers
from app.auth_middleware import require_auth, require_admin, optional_auth, get_current_user
from datetime import datetime, date
import pandas as pd
import json
import hashlib
import re
import os
import io
import csv
from collections import defaultdict

main = Blueprint('main', __name__)

@main.route('/')
@require_auth
def index():
    current_user = get_current_user()
    
    # 強制提交任何待處理的資料庫變更
    db.session.commit()
    
    today = date.today()
    
    # 獲取今日排班，排除H0和H1班別（休假），按匯入順序排列
    today_schedules = Schedule.query.join(ShiftType).filter(
        Schedule.date == today,
        ~ShiftType.code.in_(['H0', 'H1'])
    ).order_by(Schedule.import_order.asc().nullslast()).all()
    
    # 如果今日沒有排班記錄，顯示最近的排班記錄（同樣排除H0和H1）
    if not today_schedules:
        recent_schedules = Schedule.query.join(ShiftType).filter(
            ~ShiftType.code.in_(['H0', 'H1'])
        ).order_by(Schedule.date.desc(), Schedule.import_order.asc().nullslast()).limit(10).all()
        print(f"🔍 今日無排班記錄，顯示最近 {len(recent_schedules)} 筆排班記錄（排除休假）", flush=True)
        return render_template('index.html', today_schedules=recent_schedules, today=today, show_recent=True)
    
    # 調試輸出：檢查過濾結果
    print(f"🔍 今日排班記錄 (排除H0/H1): {len(today_schedules)} 筆", flush=True)
    for schedule in today_schedules:
        print(f"   {schedule.employee.name}: {schedule.shift_type.code}", flush=True)
    
    return render_template('index.html', today_schedules=today_schedules, today=today, show_recent=False)


@main.route('/calendar')
@require_auth
def calendar():
    return render_template('calendar.html')

@main.route('/api/events')
@require_auth
def get_events():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    # 如果沒有提供日期參數，使用預設範圍
    if not start_date or not end_date:
        start_date = '2024-07-01'
        end_date = '2024-07-31'
    
    try:
        schedules = Schedule.query.filter(
            Schedule.date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
            Schedule.date <= datetime.strptime(end_date, '%Y-%m-%d').date()
        ).all()
        
        events = []
        
        # 按日期分組事件，避免太多重疊
        from collections import defaultdict
        daily_events = defaultdict(list)
        
        for schedule in schedules:
            daily_events[schedule.date].append(schedule)
        
        # 為每一天創建事件
        for schedule_date, day_schedules in daily_events.items():
            if len(day_schedules) <= 3:
                # 如果當天排班人數少，顯示個別事件
                for schedule in day_schedules:
                    events.append({
                        'id': f"single_{schedule.id}",
                        'title': f"{schedule.employee.name}({schedule.shift_type.code})",
                        'start': schedule_date.isoformat(),
                        'backgroundColor': schedule.shift_type.color,
                        'borderColor': schedule.shift_type.color,
                        'textColor': '#fff',
                        'extendedProps': {
                            'employee': schedule.employee.name,
                            'shift': schedule.shift_type.name,
                            'shift_code': schedule.shift_type.code,
                            'type': 'single'
                        }
                    })
            else:
                # 如果當天排班人數多，顯示彙總事件
                shift_summary = {}
                for schedule in day_schedules:
                    shift_code = schedule.shift_type.code
                    if shift_code not in shift_summary:
                        shift_summary[shift_code] = {
                            'count': 0,
                            'color': schedule.shift_type.color,
                            'employees': []
                        }
                    shift_summary[shift_code]['count'] += 1
                    shift_summary[shift_code]['employees'].append(schedule.employee.name)
                
                # 為每個班別創建一個事件
                for shift_code, info in shift_summary.items():
                    events.append({
                        'id': f"summary_{schedule_date}_{shift_code}",
                        'title': f"{shift_code}班 ({info['count']}人)",
                        'start': schedule_date.isoformat(),
                        'backgroundColor': info['color'],
                        'borderColor': info['color'],
                        'textColor': '#fff',
                        'extendedProps': {
                            'employees': info['employees'],
                            'shift_code': shift_code,
                            'count': info['count'],
                            'type': 'summary',
                            'date': schedule_date.isoformat()
                        }
                    })
        
        return jsonify(events)
        
    except Exception as e:
        print(f"Events API 錯誤: {e}")
        return jsonify([]), 500

@main.route('/query_shift')
@require_auth
def query_shift():
    query_date = request.args.get('date')
    schedules = []
    
    # 指定的員工名單（包含空格格式）
    target_employees = [
        '賴 秉 宏', '李 惟 綱', '李 家 瑋', '王 志 忠', '顧 育 禎', 
        '胡 翊 潔', '朱 家 德', '陳 韋 如', '葛 禎', '井 康 羽', 
        '簡 芳 瑜', '梁 弘 岳', '李 佩 璇', '鄭 栢 玔', '王 文 怡'
    ]
    
    if query_date:
        try:
            target_date = datetime.strptime(query_date, '%Y-%m-%d').date()
            schedules = Schedule.query.filter(Schedule.date == target_date).all()
        except ValueError:
            query_date = None
    
    return render_template('query.html', query_date=query_date, schedules=schedules)

@main.route('/api/query_shift')
@require_auth
def api_query_shift():
    query_date = request.args.get('date')
    
    if query_date:
        target_date = datetime.strptime(query_date, '%Y-%m-%d').date()
        schedules = Schedule.query.filter(Schedule.date == target_date).all()
        
        result = []
        for schedule in schedules:
            result.append({
                'employee': schedule.employee.name,
                'shift': schedule.shift_type.name,
                'shift_code': schedule.shift_type.code,
                'color': schedule.shift_type.color
            })
        
        return jsonify(result)
    
    return jsonify([])




@main.route('/api/shift_types')
def api_shift_types():
    shift_types = ShiftType.query.all()
    result = [{'code': st.code, 'name': st.name, 'color': st.color} for st in shift_types]
    return jsonify(result)

@main.route('/api/export_ics')
def api_export_ics():
    """匯出指定員工的月度班表為ICS日曆檔案"""
    try:
        # 取得查詢參數
        search_query = request.args.get('query', '').strip()
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not search_query:
            return jsonify({'error': '請輸入員工姓名或工號'}), 400
            
        if not year or not month:
            return jsonify({'error': '請選擇年月'}), 400
            
        # 確保測試員工存在（與preview_schedule相同邏輯）
        from app.models import db
        employee_data = [
            ('賴 秉 宏', '8652'),
            ('李 惟 綱', '8312'), 
            ('李 家 瑋', '8512'),
            ('王 志 忠', '0450'),
            ('顧 育 禎', '8672'),
            ('胡 翊 潔', '8619'),
            ('朱 家 德', '8835')
        ]
        
        # 每次請求時確保員工存在（記憶體資料庫需要）
        for name, code in employee_data:
            employee = Employee.query.filter_by(name=name).first()
            if not employee:
                employee = Employee(name=name, employee_code=code)
                db.session.add(employee)
        db.session.commit()
        
        # 搜尋員工（與preview_schedule相同邏輯）
        employee = Employee.query.filter(
            Employee.name.like(f'%{search_query}%')
        ).first()
        
        # 如果模糊搜尋找不到，嘗試精確匹配
        if not employee:
            employee = Employee.query.filter_by(name=search_query).first()
        
        # 如果還是找不到，嘗試移除空格的匹配
        if not employee:
            search_no_space = search_query.replace(' ', '')
            for emp in Employee.query.all():
                if emp.name.replace(' ', '') == search_no_space:
                    employee = emp
                    break
        
        if not employee:
            return jsonify({'error': f'找不到員工: {search_query}'}), 404
            
        # 取得該員工的月度排班資料
        from datetime import date, datetime, time
        import calendar
        
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        schedules = Schedule.query.filter(
            Schedule.employee_id == employee.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date
        ).order_by(Schedule.date.asc()).all()
        
        # 生成ICS內容
        ics_content = generate_ics_content(employee, schedules, year, month)
        
        # 創建響應，避免中文檔名編碼問題
        import urllib.parse
        filename = f"{employee.name}_{year}年{month:02d}月_班表.ics"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        response = Response(
            ics_content,
            mimetype='text/calendar',
            headers={
                'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}'
            }
        )
        
        return response
        
    except Exception as e:
        print(f'ICS匯出錯誤: {e}')
        return jsonify({'error': f'匯出失敗: {str(e)}'}), 500

def generate_ics_content(employee, schedules, year, month):
    """根據Google日曆規範生成ICS檔案內容"""
    from datetime import datetime, time
    
    # 按Google日曆規範設定ICS檔案標頭
    ics_lines = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//My Calendar Generator//EN',
        'CALSCALE:GREGORIAN',
        'METHOD:PUBLISH',
        f'X-WR-CALNAME:{employee.name}的班表',
        'X-WR-TIMEZONE:Asia/Taipei'
    ]
    
    # 處理所有班別，包含休假班別
    for schedule in schedules:
        
        # 取得班別的實際工作時間
        actual_start = schedule.shift_type.start_time or time(9, 0)
        actual_end = schedule.shift_type.end_time or time(18, 0)
        
        # 按Google日曆規範：所有事件統一設為09:00-09:05
        event_start = datetime.combine(schedule.date, time(9, 0))
        event_end = datetime.combine(schedule.date, time(9, 5))
        
        # 格式化為ICS時間格式
        dtstart = event_start.strftime('%Y%m%dT%H%M%S')
        dtend = event_end.strftime('%Y%m%dT%H%M%S')
        
        # 建立班別標題，只顯示班別代碼
        summary = f'{schedule.shift_type.code}'
        
        # 按照Google日曆規範的簡化格式創建事件
        event_lines = [
            'BEGIN:VEVENT',
            f'DTSTART:{dtstart}',
            f'DTEND:{dtend}',
            f'SUMMARY:{summary}',
            'END:VEVENT'
        ]
        
        ics_lines.extend(event_lines)
    
    # ICS檔案結尾
    ics_lines.append('END:VCALENDAR')
    
    return '\r\n'.join(ics_lines)

@main.route('/api/date_range')
def api_date_range():
    """獲取資料庫中實際的日期範圍，支援多月份"""
    first_schedule = Schedule.query.order_by(Schedule.date.asc()).first()
    last_schedule = Schedule.query.order_by(Schedule.date.desc()).first()
    
    if first_schedule and last_schedule:
        # 獲取所有有資料的日期
        all_dates = [s.date.strftime('%Y-%m-%d') for s in Schedule.query.all()]
        unique_dates = sorted(list(set(all_dates)))
        
        # 統計各月份的資料
        month_stats = {}
        for schedule in Schedule.query.all():
            year_month = f'{schedule.date.year}-{schedule.date.month:02d}'
            if year_month not in month_stats:
                month_stats[year_month] = {
                    'count': 0,
                    'year': schedule.date.year,
                    'month': schedule.date.month,
                    'display': f'{schedule.date.year}年{schedule.date.month:02d}月'
                }
            month_stats[year_month]['count'] += 1
        
        # 如果跨月份，顯示範圍；如果同月份，顯示單月
        if first_schedule.date.month == last_schedule.date.month and first_schedule.date.year == last_schedule.date.year:
            month_display = first_schedule.date.strftime('%Y年%m月')
        else:
            month_display = f'{first_schedule.date.strftime("%Y年%m月")} 至 {last_schedule.date.strftime("%Y年%m月")}'
        
        return jsonify({
            'start_date': first_schedule.date.strftime('%Y-%m-%d'),
            'end_date': last_schedule.date.strftime('%Y-%m-%d'),
            'month_year': month_display,
            'year': first_schedule.date.year,
            'month': first_schedule.date.month,
            'available_dates': unique_dates,
            'available_months': month_stats
        })
    else:
        return jsonify({
            'start_date': '2025-07-01',
            'end_date': '2025-07-31',
            'month_year': '2025年07月',
            'year': 2025,
            'month': 7,
            'available_dates': [],
            'available_months': {}
        })

@main.route('/export_schedule')
@require_auth
def export_schedule():
    """一鍵匯出月度班表頁面"""
    return render_template('export.html')

@main.route('/api/preview_schedule')
def api_preview_schedule():
    """提供班表預覽資料給前端生成JPG"""
    try:
        # 取得查詢參數
        search_query = request.args.get('query', '').strip()
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not search_query:
            return jsonify({'error': '請輸入員工姓名或工號'}), 400
            
        if not year or not month:
            return jsonify({'error': '請選擇年月'}), 400
            
        # 確保測試員工存在
        from app.models import db
        employee_data = [
            ('賴 秉 宏', '8652'),
            ('李 惟 綱', '8312'), 
            ('李 家 瑋', '8512'),
            ('王 志 忠', '0450'),
            ('顧 育 禎', '8672'),
            ('胡 翊 潔', '8619'),
            ('朱 家 德', '8835')
        ]
        
        # 每次請求時確保員工存在（記憶體資料庫需要）
        for name, code in employee_data:
            employee = Employee.query.filter_by(name=name).first()
            if not employee:
                employee = Employee(name=name, employee_code=code)
                db.session.add(employee)
        db.session.commit()
        
        # 增強的員工搜尋邏輯（支援多種模糊匹配）
        employee = None
        
        # 1. 先嘗試員工代碼完全匹配
        employee = Employee.query.filter_by(employee_code=search_query).first()
        
        # 2. 嘗試員工代碼部分匹配
        if not employee:
            employee = Employee.query.filter(
                Employee.employee_code.like(f'%{search_query}%')
            ).first()
        
        # 3. 嘗試姓名完全匹配
        if not employee:
            employee = Employee.query.filter_by(name=search_query).first()
        
        # 4. 對於短查詢字串（1-2個字），直接進入多員工匹配邏輯
        if not employee and len(search_query.replace(' ', '')) <= 2:
            search_clean = search_query.replace(' ', '')
            matching_employees = []
            for emp in Employee.query.all():
                emp_name_clean = emp.name.replace(' ', '')
                # 支援輸入「李」找到所有姓李的員工
                if search_clean in emp_name_clean:
                    matching_employees.append(emp)
            
            # 如果找到多個匹配員工，返回選擇列表讓用戶選擇
            if len(matching_employees) > 1:
                matching_employees.sort(key=lambda e: e.employee_code)
                employee_choices = []
                for emp in matching_employees:
                    employee_choices.append({
                        'employee_code': emp.employee_code,
                        'name': emp.name,
                        'id': emp.id
                    })
                return jsonify({
                    'multiple_matches': True,
                    'choices': employee_choices,
                    'query': search_query
                })
            elif len(matching_employees) == 1:
                employee = matching_employees[0]
        
        # 5. 嘗試姓名模糊搜尋（包含查詢字串）- 僅限於較長的查詢字串
        if not employee and len(search_query.replace(' ', '')) > 2:
            employee = Employee.query.filter(
                Employee.name.like(f'%{search_query}%')
            ).first()
        
        # 6. 移除空格後的匹配（處理空格問題）
        if not employee:
            search_no_space = search_query.replace(' ', '')
            for emp in Employee.query.all():
                # 姓名移除空格比對
                if emp.name.replace(' ', '') == search_no_space:
                    employee = emp
                    break
                # 員工代碼移除空格比對
                if emp.employee_code.replace(' ', '') == search_no_space:
                    employee = emp
                    break
        
        if not employee:
            return jsonify({'error': f'找不到員工: {search_query}'}), 404
            
        # 取得該員工的月度排班資料
        from datetime import date
        import calendar
        
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        schedules = Schedule.query.filter(
            Schedule.employee_id == employee.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date
        ).order_by(Schedule.date.asc()).all()
        
        # 允許顯示部分資料或空白月曆，不再要求必須有完整資料
        # if not schedules:
        #     return jsonify({'error': f'{employee.name} 在 {year}年{month:02d}月 沒有排班資料'}), 404
            
        # 建立月曆結構 (使用星期日開始的傳統格式)
        calendar.setfirstweekday(calendar.SUNDAY)
        cal = calendar.monthcalendar(year, month)
        calendar.setfirstweekday(calendar.MONDAY)  # 恢復預設
        
        # 準備排班資料
        schedule_data = []
        for schedule in schedules:
            schedule_data.append({
                'date': schedule.date.isoformat(),
                'shift_code': schedule.shift_type.code,
                'shift_name': schedule.shift_type.name
            })
        
        # 準備回傳資料
        response_data = {
            'employee': {
                'name': employee.name,
                'employee_code': employee.employee_code
            },
            'year': year,
            'month': month,
            'schedules': schedule_data,
            'calendar': cal
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f'預覽API錯誤: {e}')
        return jsonify({'error': str(e)}), 500

@main.route('/api/preview_schedule_by_id')
def api_preview_schedule_by_id():
    """根據員工ID直接預覽班表（用於用戶選擇特定員工後）"""
    try:
        employee_id = request.args.get('employee_id')
        year = int(request.args.get('year'))
        month = int(request.args.get('month'))
        
        if not employee_id:
            return jsonify({'error': '缺少員工ID'}), 400
        
        # 直接根據ID查找員工
        employee = Employee.query.get(employee_id)
        if not employee:
            return jsonify({'error': f'找不到員工ID: {employee_id}'}), 404
        
        # 取得該員工的月度排班資料
        from datetime import date
        import calendar
        
        # 生成指定月份的所有日期
        year_month = date(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]
        
        schedules = Schedule.query.filter(
            Schedule.employee_id == employee.id,
            db.extract('year', Schedule.date) == year,
            db.extract('month', Schedule.date) == month
        ).all()
        
        # 將排班資料轉換為字典格式，方便前端處理
        schedule_dict = {schedule.date.day: schedule for schedule in schedules}
        
        # 生成月曆結構 (使用星期日開始的傳統格式)
        calendar.setfirstweekday(calendar.SUNDAY)
        cal = calendar.monthcalendar(year, month)
        calendar.setfirstweekday(calendar.MONDAY)  # 恢復預設
        
        # 準備返回的排班資料
        schedule_data = []
        for schedule in schedules:
            # 透過關聯獲取班別類型資料
            shift_type = schedule.shift_type
            schedule_data.append({
                'date': schedule.date.strftime('%Y-%m-%d'),
                'shift_code': shift_type.code if shift_type else 'UNKNOWN',
                'shift_name': shift_type.name if shift_type else 'Unknown Shift'
            })
        
        return jsonify({
            'employee': {
                'name': employee.name,
                'employee_code': employee.employee_code
            },
            'year': year,
            'month': month,
            'calendar': cal,
            'schedules': schedule_data
        })
        
    except Exception as e:
        print(f'按ID預覽API錯誤: {e}')
        return jsonify({'error': str(e)}), 500

@main.route('/api/export_monthly_schedule')
def api_export_monthly_schedule():
    """匯出指定員工的月度班表為Excel"""
    try:
        # 取得查詢參數
        search_query = request.args.get('query', '').strip()
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not search_query:
            return jsonify({'error': '請輸入員工姓名或工號'}), 400
            
        if not year or not month:
            return jsonify({'error': '請選擇年月'}), 400
            
        # 搜尋員工（支援姓名或員工代碼）
        employee = Employee.query.filter(
            (Employee.name.like(f'%{search_query}%')) |
            (Employee.employee_code.like(f'%{search_query}%'))
        ).first()
        
        if not employee:
            return jsonify({'error': f'找不到員工: {search_query}'}), 404
            
        # 取得該員工的月度排班資料
        from datetime import date
        start_date = date(year, month, 1)
        
        # 計算月份的最後一天
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        schedules = Schedule.query.filter(
            Schedule.employee_id == employee.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date
        ).order_by(Schedule.date.asc()).all()
        
        if not schedules:
            return jsonify({'error': f'{employee.name} 在 {year}年{month:02d}月 沒有排班資料'}), 404
            
        # 建立月曆式Excel匯出
        import pandas as pd
        from io import BytesIO
        import calendar
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # 建立工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '月度班表'
        
        # 設定標題
        ws['A1'] = f'{employee.name} 個人班表'
        ws['A2'] = f'工號: {employee.employee_code}'
        ws['A3'] = f'{year}年{month:02d}月'
        
        # 合併標題儲存格
        ws.merge_cells('A1:H1')
        ws.merge_cells('A2:H2')
        ws.merge_cells('A3:H3')
        
        # 設定標題樣式
        title_font = Font(name='微軟正黑體', size=16, bold=True)
        subtitle_font = Font(name='微軟正黑體', size=12)
        ws['A1'].font = title_font
        ws['A2'].font = subtitle_font
        ws['A3'].font = subtitle_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # 建立班表資料字典
        schedule_dict = {}
        for schedule in schedules:
            schedule_dict[schedule.date.day] = schedule.shift_type.code
        
        # 取得該月的日曆資訊 (使用星期日開始的傳統格式)
        calendar.setfirstweekday(calendar.SUNDAY)
        cal = calendar.monthcalendar(year, month)
        calendar.setfirstweekday(calendar.MONDAY)  # 恢復預設
        weekdays = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
        
        # 從第5行開始建立日曆表格
        start_row = 5
        
        # 寫入星期標題
        for col, weekday in enumerate(weekdays, 1):
            cell = ws.cell(row=start_row, column=col, value=weekday)
            cell.font = Font(name='微軟正黑體', size=12, bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # 設定邊框樣式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填入日期和班別
        for week_num, week in enumerate(cal):
            row = start_row + 1 + week_num
            for day_num, day in enumerate(week, 1):
                if day == 0:  # 空白日期
                    cell = ws.cell(row=row, column=day_num, value='')
                else:
                    # 日期
                    date_str = str(day)
                    # 班別（如果有的話）
                    shift_code = schedule_dict.get(day, '')
                    
                    if shift_code:
                        cell_value = f'{date_str}\n{shift_code}'
                        # 設定班別背景色
                        if 'H' in shift_code:  # 休假
                            fill_color = 'FFE6CC'  # 淺橙色
                        elif 'P1' in shift_code:  # P1班
                            fill_color = 'E2EFDA'  # 淺綠色
                        elif 'P2' in shift_code:  # P2班
                            fill_color = 'FFF2CC'  # 淺黃色
                        elif 'P3' in shift_code:  # P3班
                            fill_color = 'FCE4D6'  # 淺橙色
                        elif 'P4' in shift_code:  # P4班
                            fill_color = 'F4CCCC'  # 淺紅色
                        elif 'FC' in shift_code:  # FC班
                            fill_color = 'D0E0E3'  # 淺藍色
                        else:
                            fill_color = 'F2F2F2'  # 淺灰色
                        
                        cell = ws.cell(row=row, column=day_num, value=cell_value)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        cell = ws.cell(row=row, column=day_num, value=date_str)
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    
                    cell.font = Font(name='微軟正黑體', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 添加邊框
                ws.cell(row=row, column=day_num).border = thin_border
        
        # 為星期標題行也添加邊框
        for col in range(1, 8):
            ws.cell(row=start_row, column=col).border = thin_border
        
        # 設定欄位寬度
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        # 設定行高
        for row in range(start_row + 1, start_row + 1 + len(cal)):
            ws.row_dimensions[row].height = 40
        
        # 添加圖例說明
        legend_start_row = start_row + len(cal) + 2
        ws.cell(row=legend_start_row, column=1, value='班別說明:').font = Font(name='微軟正黑體', size=12, bold=True)
        
        # 取得該員工使用的班別類型
        used_shift_types = set()
        for schedule in schedules:
            used_shift_types.add((schedule.shift_type.code, schedule.shift_type.color))
        
        legend_row = legend_start_row + 1
        for i, (code, color) in enumerate(sorted(used_shift_types)):
            legend_cell = ws.cell(row=legend_row + i, column=1, value=f'{code}: {code}')
            legend_cell.font = Font(name='微軟正黑體', size=10)
            # 設定背景色與班別顏色相同
            if color and color.startswith('#'):
                # 移除#號並轉為大寫
                hex_color = color[1:].upper()
                legend_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
            
        output.seek(0)
        
        # 準備檔案名稱
        filename = f'{employee.name}_{year}年{month:02d}月_班表.xlsx'
        
        # 返回檔案下載
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f'匯出錯誤: {e}')
        return jsonify({'error': str(e)}), 500

@main.route('/upload_excel', methods=['GET', 'POST'])
@require_admin
def upload_excel():
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                return render_template('upload.html', error='請選擇檔案')
            
            file = request.files['file']
            if file.filename == '':
                return render_template('upload.html', error='請選擇檔案')
            
            if not file.filename.endswith('.csv'):
                return render_template('upload.html', error='只支援 CSV 格式檔案')
            
            print(f'開始處理檔案: {file.filename}')
            
            # 嘗試不同的方法讀取 Excel 檔案
            df = None
            for header_row in [0, 1, 2, 3, 4, 5]:  # 嘗試前6行作為標題行
                try:
                    temp_df = pd.read_excel(file, header=header_row)
                    print(f'嘗試標題行 {header_row}: {list(temp_df.columns)[:10]}')  # 只顯示前10個欄位
                    
                    # 檢查是否包含我們需要的欄位
                    columns_str = '|'.join(str(col).lower() for col in temp_df.columns)
                    if any(keyword in columns_str for keyword in ['姓名', '員工', '日期', '班', '時間']):
                        df = temp_df
                        print(f'✅ 在第 {header_row} 行找到有效標題')
                        break
                except Exception as e:
                    print(f'標題行 {header_row} 讀取失敗: {e}')
                    continue
            
            if df is None:
                # 如果沒有找到有效標題，使用第一行
                df = pd.read_excel(file)
                print('⚠️ 未找到明確標題行，使用預設格式')
            
            print(f'最終讀取到 {len(df)} 行數據')
            print(f'最終欄位: {list(df.columns)}')
            
            processed_count = 0
            error_count = 0
            
            # 偵測橫向班表格式
            for index, row in df.iterrows():
                try:
                    # 尋找員工姓名（通常在第3或第4欄）
                    employee_name = None
                    for col_idx in [3, 2, 1, 0]:  # 檢查可能的姓名欄位
                        col_name = f'Unnamed: {col_idx}' if col_idx > 0 else 'Unnamed: 0'
                        if col_name in row and not pd.isna(row[col_name]):
                            name_value = str(row[col_name]).strip()
                            # 檢查是否看起來像姓名（包含中文字且不是數字或班別代碼）
                            if name_value and not name_value.isdigit() and len(name_value) >= 2 and '（異動後）' not in name_value:
                                # 進一步檢查是否為有效姓名格式
                                if any('\u4e00' <= char <= '\u9fff' for char in name_value):  # 包含中文
                                    employee_name = name_value
                                    print(f'找到員工姓名: {employee_name} (第 {index+1} 行)')
                                    break
                    
                    if not employee_name:
                        continue
                    
                    # 處理橫向班表：每個欄位代表一天的班別
                    current_date = datetime(2025, 7, 1).date()  # 114年7月 = 2025年7月
                    
                    for col_name, shift_value in row.items():
                        if col_name.startswith('Unnamed:') and not pd.isna(shift_value):
                            shift_code = str(shift_value).strip()
                            print(f'  檢查欄位 {col_name}: 值={shift_value} -> 班別代碼={shift_code}')
                            
                            # 過濾掉非班別代碼的值
                            if shift_code and shift_code not in ['nan', employee_name] and len(shift_code) <= 10:
                                # 檢查是否為有效的班別代碼 - 放寬條件
                                print(f'  檢查班別代碼 {shift_code} 是否有效...')
                                # 更寬鬆的班別代碼驗證：只要不是純數字且長度合理
                                if (not shift_code.isdigit() and 
                                    len(shift_code) >= 1 and 
                                    shift_code not in ['nan', 'NaN', 'None', ''] and
                                    not shift_code.replace('.', '').isdigit()):
                                    print(f'  ✅ 找到有效班別: {shift_code}')
                                    
                                    # 查詢現有員工記錄
                                    existing_employee = Employee.query.filter_by(name=employee_name).first()
                                    if not existing_employee:
                                        # 生成唯一的員工代碼
                                        existing_count = Employee.query.count()
                                        employee_code = f"EMP_{existing_count+1:03d}"
                                        
                                        # 確保代碼唯一性
                                        while Employee.query.filter_by(employee_code=employee_code).first():
                                            existing_count += 1
                                            employee_code = f"EMP_{existing_count+1:03d}"
                                        
                                        new_employee = Employee(name=employee_name, employee_code=employee_code)
                                        db.session.add(new_employee)
                                        try:
                                            db.session.commit()
                                        except Exception as e:
                                            db.session.rollback()
                                            existing_employee = Employee.query.filter_by(name=employee_name).first()
                                            if not existing_employee:
                                                raise e
                                    else:
                                        new_employee = existing_employee
                                    
                                    employee = new_employee if 'new_employee' in locals() else existing_employee
                                    
                                    # 處理班別類型
                                    shift_type = ShiftType.query.filter_by(code=shift_code).first()
                                    if not shift_type:
                                        shift_name, start_time, end_time, color = get_shift_info(shift_code)
                                        shift_type = ShiftType(
                                            code=shift_code, 
                                            name=shift_name, 
                                            start_time=start_time, 
                                            end_time=end_time,
                                            color=color
                                        )
                                        db.session.add(shift_type)
                                    
                                    # 計算對應的日期（依欄位位置推算）
                                    col_num = int(col_name.split(': ')[1]) if ': ' in col_name else 0
                                    day_offset = col_num - 4  # 假設第4欄開始是7/1
                                    if day_offset >= 0 and day_offset < 31:  # 7月有31天
                                        schedule_date = datetime(2025, 7, day_offset + 1).date()
                                        
                                        # 檢查是否已存在記錄
                                        existing_schedule = Schedule.query.filter_by(
                                            date=schedule_date,
                                            employee=employee
                                        ).first()
                                        
                                        if not existing_schedule:
                                            schedule = Schedule(
                                                date=schedule_date,
                                                employee=employee,
                                                shift_type=shift_type
                                            )
                                            db.session.add(schedule)
                                            processed_count += 1
                                            print(f'處理: {employee_name} - {schedule_date} - {shift_code}')
                    
                except Exception as row_error:
                    error_count += 1
                    print(f'第 {index+1} 行處理錯誤: {row_error}')
                    continue
            
            db.session.commit()
            success_message = f'匯入完成！處理了 {processed_count} 筆記錄，{error_count} 筆錯誤'
            print(success_message)
            return render_template('upload.html', success=success_message)
            
        except Exception as e:
            db.session.rollback()
            error_message = f'匯入失敗: {str(e)}'
            print(error_message)
            return render_template('upload.html', error=error_message)
    
    return render_template('upload.html')

def get_shift_info(shift_code):
    """根據班別代碼返回班別資訊，支援複合格式和未排班狀態"""
    from datetime import time
    
    # 特殊狀態：未排班
    if shift_code == 'UNASSIGNED':
        return '未排班', time(0, 0), time(0, 0), '#9ca3af'
    
    # 休假相關
    if shift_code in ['H0', 'H1', 'H2']:
        return f'休假({shift_code})', time(0, 0), time(0, 0), '#6c757d'
    
    # 正常班別
    elif shift_code == 'FC':
        return 'FC班', time(9, 0), time(18, 0), '#007bff'
    elif shift_code == 'FX':
        return 'FX班', time(10, 0), time(19, 0), '#495057'
    
    # 處理複合班別（如 P1c/工程, P4p/ME 等）
    base_code = shift_code.split('/')[0] if '/' in shift_code else shift_code
    suffix = shift_code.split('/')[1] if '/' in shift_code else ''
    
    # P班系列（根據基礎代碼判斷時間）
    if base_code.startswith('P1'):
        name = f'{shift_code}班' if '/' in shift_code else f'P1班({shift_code})'
        return name, time(8, 0), time(17, 0), '#28a745'
    elif base_code.startswith('P2'):
        name = f'{shift_code}班' if '/' in shift_code else f'P2班({shift_code})'
        return name, time(14, 0), time(23, 0), '#ffc107'
    elif base_code.startswith('P3'):
        name = f'{shift_code}班' if '/' in shift_code else f'P3班({shift_code})'
        return name, time(17, 0), time(2, 0), '#fd7e14'
    elif base_code.startswith('P4'):
        name = f'{shift_code}班' if '/' in shift_code else f'P4班({shift_code})'
        return name, time(20, 0), time(5, 0), '#dc3545'
    elif base_code.startswith('P5'):
        name = f'{shift_code}班' if '/' in shift_code else f'P5班({shift_code})'
        return name, time(22, 0), time(7, 0), '#6f42c1'
    
    # 其他班別系列
    elif base_code.startswith('N'):
        return f'{shift_code}班', time(19, 0), time(4, 0), '#20c997'
    elif base_code.startswith('E'):
        return f'{shift_code}班', time(7, 0), time(16, 0), '#17a2b8'
    elif base_code.startswith('C'):
        return f'{shift_code}班', time(16, 0), time(1, 0), '#e83e8c'
    elif base_code.startswith('R'):
        return f'{shift_code}班', time(12, 0), time(21, 0), '#6610f2'
    elif shift_code in ['NT', 'CH']:
        return f'{shift_code}班', time(9, 0), time(18, 0), '#343a40'
    
    # 特殊格式處理（如 FC/前製）
    elif base_code == 'FC' and suffix:
        return f'FC班/{suffix}', time(9, 0), time(18, 0), '#007bff'
    
    # 其他未知班別（強制匯入時可能產生）
    else:
        return f'{shift_code}班', time(9, 0), time(18, 0), '#868e96'

@main.route('/employee/<int:employee_id>/schedule')
def employee_schedule(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    schedules = Schedule.query.filter_by(employee_id=employee_id).order_by(Schedule.date.desc()).all()
    today = date.today()
    
    return render_template('employee_schedule.html', employee=employee, schedules=schedules, today=today)

# ===== 管理員功能 =====

@main.route('/api/admin/pending-users')
@require_admin
def get_pending_users():
    """獲取待審核用戶列表"""
    try:
        pending_users = User.query.filter_by(status='pending').order_by(User.created_at.desc()).all()
        
        result = []
        for user in pending_users:
            result.append({
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'status': user.status,
                'created_at': user.created_at.isoformat()
            })
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f'獲取待審核用戶錯誤: {e}')
        return jsonify({'success': False, 'message': '獲取用戶列表失敗'}), 500

@main.route('/api/admin/approve-user', methods=['POST'])
@require_admin
def approve_user():
    """審核用戶"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': '請提供審核資料'}), 400
        
        # 支援兩種參數格式：userId 和 user_id
        user_id = data.get('userId') or data.get('user_id')
        action = data.get('action')  # 'approve' or 'reject'
        
        # 確保user_id是整數類型
        try:
            if user_id:
                user_id = int(user_id)
        except (ValueError, TypeError):
            return jsonify({
                'success': False, 
                'message': f'用戶ID必須是數字，收到: {user_id}'
            }), 400
        
        if not user_id or action not in ['approve', 'reject']:
            return jsonify({
                'success': False, 
                'message': f'無效的審核參數 - user_id: {user_id}, action: {action}'
            }), 400
        
        user = User.query.get(user_id)
        if not user:
            # 增加診斷信息：查看所有用戶
            all_users = User.query.all()
            print(f'🔍 找不到用戶ID {user_id}，現有用戶:')
            for u in all_users:
                print(f'   ID: {u.id}, 用戶名: {u.username}, 狀態: {u.status}')
            return jsonify({'success': False, 'message': f'用戶不存在 (ID: {user_id})'}), 404
        
        if user.status != 'pending':
            return jsonify({'success': False, 'message': '用戶狀態不是待審核'}), 400
        
        # 更新用戶狀態
        if action == 'approve':
            user.status = 'approved'
            message = f'用戶 {user.name} 已審核通過'
        else:
            user.status = 'rejected'
            message = f'用戶 {user.name} 已被拒絕'
        
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        db.session.rollback()
        print(f'審核用戶錯誤: {e}')
        return jsonify({'success': False, 'message': '審核失敗'}), 500

@main.route('/api/admin/users')
@require_admin
def get_all_users():
    """獲取所有用戶列表"""
    try:
        users = User.query.order_by(User.created_at.desc()).all()
        
        result = []
        for user in users:
            result.append({
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'role': user.role,
                'status': user.status,
                'failed_attempts': user.failed_attempts,
                'is_locked': user.is_locked(),
                'created_at': user.created_at.isoformat(),
                'updated_at': user.updated_at.isoformat() if user.updated_at else None
            })
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f'獲取用戶列表錯誤: {e}')
        return jsonify({'success': False, 'message': '獲取用戶列表失敗'}), 500

@main.route('/api/admin/user/<int:user_id>/status', methods=['PUT'])
@require_admin
def update_user_status(user_id):
    """更新用戶狀態"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': '請提供狀態資料'}), 400
        
        new_status = data.get('status')
        
        if new_status not in ['pending', 'approved', 'rejected', 'disabled']:
            return jsonify({'success': False, 'message': '無效的狀態值'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用戶不存在'}), 404
        
        # 防止管理員修改自己的狀態
        current_user = get_current_user()
        if user.id == current_user.id:
            return jsonify({'success': False, 'message': '不能修改自己的狀態'}), 400
        
        user.status = new_status
        user.updated_at = datetime.utcnow()
        
        # 如果啟用用戶，重置失敗次數
        if new_status == 'approved':
            user.reset_failed_attempts()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'用戶狀態已更新為 {new_status}'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f'更新用戶狀態錯誤: {e}')
        return jsonify({'success': False, 'message': '更新失敗'}), 500

@main.route('/api/admin/user/<int:user_id>/unlock', methods=['POST'])
@require_admin
def unlock_user(user_id):
    """解鎖用戶帳號"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用戶不存在'}), 404
        
        user.reset_failed_attempts()
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'用戶 {user.name} 已解鎖'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f'解鎖用戶錯誤: {e}')
        return jsonify({'success': False, 'message': '解鎖失敗'}), 500

@main.route('/api/admin/user/<int:user_id>', methods=['DELETE'])
@require_admin
def delete_user(user_id):
    """刪除用戶"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用戶不存在'}), 404
        
        # 檢查是否為系統管理員
        if user.role == 'admin':
            return jsonify({'success': False, 'message': '不能刪除管理員帳號'}), 403
        
        # 檢查是否為當前登入用戶
        current_user_id = session.get('user_id')
        if user.id == current_user_id:
            return jsonify({'success': False, 'message': '不能刪除當前登入的帳號'}), 403
        
        # 備份用戶資料以便日誌記錄
        deleted_user_name = user.name
        deleted_user_username = user.username
        
        # 刪除用戶相關的班表記錄
        from app.models import Schedule
        Schedule.query.filter_by(employee_id=user.id).delete()
        
        # 刪除用戶
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'用戶 {deleted_user_name} ({deleted_user_username}) 已刪除'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f'刪除用戶錯誤: {e}')
        return jsonify({'success': False, 'message': '刪除失敗'}), 500

@main.route('/admin')
@require_admin
def admin_page():
    """管理員控制台頁面"""
    return render_template('admin.html')

# ===== 新版Excel匯入驗證系統 v1.0 =====

@main.route('/upload_new', methods=['GET', 'POST'])
@require_admin
def upload_new():
    """新版Excel匯入驗證系統"""
    if request.method == 'POST':
        return handle_excel_upload()
    
    # GET請求：顯示匯入頁面
    recent_imports = ImportLog.query.order_by(ImportLog.import_time.desc()).limit(5).all()
    
    # 添加當前班表統計資料
    from datetime import date, datetime, timedelta
    today = date.today()
    
    # 統計當前資料庫中的班表資料
    total_schedules = Schedule.query.count()
    total_employees = Employee.query.count()
    
    # 查找最近的排班資料範圍
    latest_schedule = Schedule.query.order_by(Schedule.date.desc()).first()
    earliest_schedule = Schedule.query.order_by(Schedule.date.asc()).first()
    
    # 今日排班統計
    today_schedules_count = Schedule.query.filter(Schedule.date == today).count()
    
    # 本月排班統計
    month_start = date(today.year, today.month, 1)
    if today.month == 12:
        month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    month_schedules_count = Schedule.query.filter(
        Schedule.date >= month_start,
        Schedule.date <= month_end
    ).count()
    
    current_stats = {
        'total_schedules': total_schedules,
        'total_employees': total_employees,
        'today_schedules': today_schedules_count,
        'month_schedules': month_schedules_count,
        'latest_date': latest_schedule.date if latest_schedule else None,
        'earliest_date': earliest_schedule.date if earliest_schedule else None,
        'today': today
    }
    
    return render_template('upload_new.html', 
                         recent_imports=recent_imports, 
                         current_stats=current_stats)

def handle_excel_upload():
    """處理Excel檔案上傳和驗證"""
    try:
        # 1. 檔案驗證
        if 'file' not in request.files:
            flash('請選擇檔案', 'error')
            return redirect(url_for('main.upload_new'))
        
        file = request.files['file']
        target_group = request.form.get('target_group')
        skip_invalid = request.form.get('skip_invalid') == 'on'
        
        if file.filename == '':
            flash('請選擇檔案', 'error')
            return redirect(url_for('main.upload_new'))
        
        if not target_group:
            flash('請選擇人員群組', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 自動從檔案名稱判斷資料版本
        data_version = auto_detect_version(file.filename)
        
        if not file.filename.endswith('.csv'):
            flash('只支援 CSV 格式檔案', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 2. 讀取白名單配置
        whitelist_data = load_whitelist()
        if not whitelist_data:
            flash('無法載入白名單配置', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 3. 讀取和解析Excel
        df = read_excel_file(file)
        if df is None:
            flash('無法讀取Excel檔案，請檢查檔案格式', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 4. 資料預覽（前50筆）
        preview_data = create_preview_data_v11(df, target_group, whitelist_data)
        
        # 5. 資料驗證
        validation_result = validate_excel_data_v11(df, data_version, file.filename, target_group, whitelist_data)
        
        # 6. 為新的匯入方式準備base64編碼的CSV資料
        import base64
        csv_string = df.to_csv(index=False)
        csv_base64 = base64.b64encode(csv_string.encode('utf-8')).decode('utf-8')
        validation_result['csv_data'] = csv_base64
        validation_result['target_group'] = target_group
        validation_result['filename'] = file.filename
        
        # 7. 顯示結果頁面
        recent_imports = ImportLog.query.order_by(ImportLog.import_time.desc()).limit(5).all()
        
        return render_template('upload_new.html', 
                             preview_data=preview_data,
                             validation_result=validation_result,
                             recent_imports=recent_imports)
        
    except Exception as e:
        flash(f'處理檔案時發生錯誤: {str(e)}', 'error')
        return redirect(url_for('main.upload_new'))

def read_excel_file(file):
    """讀取CSV檔案並返回DataFrame"""
    return read_csv_file(file)

def read_csv_file(file):
    """讀取CSV檔案並返回DataFrame"""
    try:
        # 嘗試不同的編碼
        encodings = ['utf-8', 'utf-8-sig', 'big5', 'gbk', 'cp950']
        
        for encoding in encodings:
            try:
                # 重置檔案指針
                file.seek(0)
                
                # 嘗試讀取CSV
                df = pd.read_csv(file, encoding=encoding)
                
                # 檢查是否包含必要欄位
                columns_str = '|'.join(str(col).lower() for col in df.columns)
                if any(keyword in columns_str for keyword in ['姓名', '員工', '日期', '班']):
                    print(f'✅ CSV檔案使用 {encoding} 編碼讀取成功')
                    return df
                    
                # 如果第一行不是標題，嘗試跳過前幾行
                for skip_rows in [1, 2, 3, 4]:
                    try:
                        file.seek(0)
                        df_skip = pd.read_csv(file, encoding=encoding, skiprows=skip_rows)
                        columns_str = '|'.join(str(col).lower() for col in df_skip.columns)
                        if any(keyword in columns_str for keyword in ['姓名', '員工', '日期', '班']):
                            print(f'✅ CSV檔案使用 {encoding} 編碼，跳過 {skip_rows} 行讀取成功')
                            return df_skip
                    except:
                        continue
                
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f'使用 {encoding} 編碼讀取CSV失敗: {e}')
                continue
        
        # 如果所有編碼都失敗，使用預設UTF-8
        file.seek(0)
        df = pd.read_csv(file, encoding='utf-8')
        print('⚠️ 使用預設UTF-8編碼讀取CSV')
        return df
        
    except Exception as e:
        print(f'讀取CSV檔案錯誤: {e}')
        return None


def create_preview_data(df):
    """創建資料預覽（前10筆）"""
    preview_data = []
    
    # 嘗試識別欄位
    name_col, date_col, shift_col = identify_columns(df)
    
    if not all([name_col, date_col, shift_col]):
        # 如果無法識別欄位，顯示原始資料
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            preview_data.append({
                '姓名': str(row.iloc[0]) if len(row) > 0 else '',
                '日期': str(row.iloc[1]) if len(row) > 1 else '',
                '班別': str(row.iloc[2]) if len(row) > 2 else ''
            })
    else:
        # 使用識別的欄位
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            preview_data.append({
                '姓名': str(row[name_col]) if pd.notna(row[name_col]) else '',
                '日期': str(row[date_col]) if pd.notna(row[date_col]) else '',
                '班別': str(row[shift_col]) if pd.notna(row[shift_col]) else ''
            })
    
    return preview_data

def identify_columns(df):
    """智能識別CSV格式並提取欄位 - 支援3欄位和5欄位格式"""
    # 初始化欄位變數
    name_col = None
    date_col = None
    shift_col = None
    employee_code_col = None
    year_month_col = None
    day_col = None
    
    # 調試輸出
    print(f"🔍 CSV欄位: {list(df.columns)}")
    print(f"🔍 欄位數量: {len(df.columns)}")
    print(f"🔍 第一行資料: {list(df.iloc[0]) if len(df) > 0 else 'N/A'}")
    
    # 判斷CSV格式類型
    csv_format = detect_csv_format(df)
    print(f"🎯 偵測到CSV格式: {csv_format}")
    
    if csv_format == "5_column":
        # 5欄位格式: 姓名, 員工代碼, 年月, 日期, 班別
        return identify_5_column_format(df)
    elif csv_format == "3_column":
        # 3欄位格式: 姓名, 日期, 班別 (原有格式)
        return identify_3_column_format(df)
    else:
        # 嘗試通用識別
        return identify_columns_generic(df)

def detect_csv_format(df):
    """偵測CSV格式類型"""
    columns = list(df.columns)
    col_count = len(columns)
    
    # 檢查5欄位格式特徵
    if col_count >= 5:
        col_names_lower = [str(col).lower().strip() for col in columns]
        
        # 尋找5欄位格式的關鍵欄位
        has_employee_code = any('員工代碼' in col or '代碼' in col or 'code' in col.lower() for col in col_names_lower)
        has_year_month = any('年月' in col for col in col_names_lower)
        has_day = any(col in ['日期', 'day', '日'] for col in col_names_lower)
        
        if has_employee_code or has_year_month:
            return "5_column"
    
    # 檢查3欄位格式特徵
    if col_count == 3:
        col_names_lower = [str(col).lower().strip() for col in columns]
        has_name = any(keyword in col for col in col_names_lower for keyword in ['姓名', 'name', '員工'])
        has_date = any(keyword in col for col in col_names_lower for keyword in ['日期', 'date'])
        has_shift = any(keyword in col for col in col_names_lower for keyword in ['班別', 'shift', '班次'])
        
        if has_name and has_date and has_shift:
            return "3_column"
    
    return "unknown"

def identify_5_column_format(df):
    """識別5欄位格式: 姓名, 員工代碼, 年月, 日期, 班別"""
    name_col = None
    employee_code_col = None
    year_month_col = None
    day_col = None
    shift_col = None
    
    for col in df.columns:
        col_str = str(col).lower().strip()
        
        # 姓名欄位
        if col_str in ['姓名', '員工姓名', 'name', '名字'] and not name_col:
            name_col = col
            print(f"✅ 找到姓名欄位: {col}")
        
        # 員工代碼欄位
        elif ('員工代碼' in col_str or '代碼' in col_str or 'code' in col_str) and not employee_code_col:
            employee_code_col = col
            print(f"✅ 找到員工代碼欄位: {col}")
        
        # 年月欄位
        elif '年月' in col_str and not year_month_col:
            year_month_col = col
            print(f"✅ 找到年月欄位: {col}")
        
        # 日期欄位 (單純的日)
        elif col_str in ['日期', 'day', '日'] and not day_col:
            day_col = col
            print(f"✅ 找到日期欄位: {col}")
        
        # 班別欄位
        elif col_str in ['班別', '班次', 'shift', '排班'] and not shift_col:
            shift_col = col
            print(f"✅ 找到班別欄位: {col}")
    
    print(f"🎯 5欄位格式識別結果: 姓名={name_col}, 員工代碼={employee_code_col}, 年月={year_month_col}, 日期={day_col}, 班別={shift_col}")
    
    # 返回包含格式資訊的結果
    return {
        'format': '5_column',
        'name_col': name_col,
        'employee_code_col': employee_code_col,
        'year_month_col': year_month_col,
        'day_col': day_col,
        'shift_col': shift_col,
        'date_col': None  # 5欄位格式中需要組合年月和日期
    }

def identify_3_column_format(df):
    """識別3欄位格式: 姓名, 日期, 班別（原有格式）"""
    name_col = None
    date_col = None
    shift_col = None
    
    # 優先處理直式格式 - 檢查column headers
    for col in df.columns:
        col_str = str(col).lower().strip()
        
        # 姓名欄位識別
        if col_str in ['姓名', '員工姓名', 'name', '名字'] and not name_col:
            name_col = col
            print(f"✅ 找到姓名欄位: {col}")
        
        # 日期欄位識別
        elif col_str in ['日期', 'date', '時間', 'datetime'] and not date_col:
            date_col = col
            print(f"✅ 找到日期欄位: {col}")
        
        # 班別欄位識別
        elif col_str in ['班別', '班次', 'shift', '排班'] and not shift_col:
            shift_col = col
            print(f"✅ 找到班別欄位: {col}")
    
    # 如果直式格式識別失敗，嘗試橫式格式（數字欄位）
    if not date_col:
        numeric_cols = [col for col in df.columns if str(col).strip().isdigit()]
        if numeric_cols:
            date_col = numeric_cols[0]  # 取第一個數字欄位作為日期
            print(f"⚡ 使用數字欄位作為日期（橫式格式）: {date_col}")
    
    print(f"🎯 3欄位格式識別結果: 姓名={name_col}, 日期={date_col}, 班別={shift_col}")
    
    # 返回與原有格式兼容的結果
    return {
        'format': '3_column',
        'name_col': name_col,
        'date_col': date_col,  
        'shift_col': shift_col,
        'employee_code_col': None,
        'year_month_col': None,
        'day_col': None
    }

def identify_columns_generic(df):
    """通用欄位識別（向後兼容）"""
    name_col = None
    date_col = None  
    shift_col = None
    
    for col in df.columns:
        col_str = str(col).lower().strip()
        
        if col_str in ['姓名', '員工姓名', 'name', '名字'] and not name_col:
            name_col = col
        elif col_str in ['日期', 'date', '時間', 'datetime'] and not date_col:
            date_col = col
        elif col_str in ['班別', '班次', 'shift', '排班'] and not shift_col:
            shift_col = col
    
    # 如果仍未找到必要欄位，檢查第一行資料是否為欄位名稱
    if not all([name_col, date_col, shift_col]) and len(df) > 0:
        first_row = df.iloc[0]
        for idx, cell in enumerate(first_row):
            cell_str = str(cell).strip()
            if cell_str == '姓名' and not name_col:
                name_col = df.columns[idx]
    
    return {
        'format': '3_column',  # 默認為3欄位兼容格式
        'name_col': name_col,
        'date_col': date_col,
        'shift_col': shift_col,
        'employee_code_col': None,
        'year_month_col': None,
        'day_col': None
    }

def combine_date_from_5_column(year_month_value, day_value):
    """從5欄位格式的年月和日期組合完整日期"""
    try:
        if pd.isna(year_month_value) or pd.isna(day_value):
            return None
            
        year_month_str = str(year_month_value).strip()
        day_str = str(day_value).strip()
        
        # 處理年月格式 (例如: "2024-10", "2024/10", "202410", "114/08"民國年)
        year = None
        month = None
        
        if '-' in year_month_str:
            year, month = year_month_str.split('-', 1)
        elif '/' in year_month_str:
            year, month = year_month_str.split('/', 1)
        elif len(year_month_str) == 6 and year_month_str.isdigit():  # 202410
            year = year_month_str[:4]
            month = year_month_str[4:]
        else:
            print(f"⚠️ 無法解析年月格式: {year_month_str}")
            return None
        
        # 處理年份格式
        year = year.strip()
        month = month.strip()
        
        # 檢查是否為民國年格式 (例如: 114 = 民國114年 = 西元2025年)
        if len(year) <= 3 and year.isdigit():
            year_int = int(year)
            if year_int > 10 and year_int < 200:  # 假設是民國年
                year = str(year_int + 1911)  # 轉換為西元年
                print(f"🗓️ 偵測到民國年，轉換: 民國{year_int}年 → 西元{year}年")
        
        # 處理日期 (去除前導零)
        day = day_str.lstrip('0') or '1'  # 如果全是0，設為1
        
        # 驗證年份格式
        if not year.isdigit() or len(year) != 4:
            print(f"⚠️ 年份格式錯誤: {year}")
            return None
        
        # 驗證月份格式
        if not month.isdigit() or int(month) < 1 or int(month) > 12:
            print(f"⚠️ 月份格式錯誤: {month}")
            return None
        
        # 組合完整日期
        full_date_str = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # 驗證日期有效性
        combined_date = datetime.strptime(full_date_str, '%Y-%m-%d').date()
        print(f"📅 日期組合成功: {year_month_str} + {day_str} = {combined_date}")
        
        return combined_date
        
    except Exception as e:
        print(f"❌ 日期組合失敗: {year_month_str} + {day_str}, 錯誤: {e}")
        return None

def get_date_value_enhanced(row, columns_info):
    """根據格式獲取日期值 - 支援3欄位和5欄位格式"""
    csv_format = columns_info.get('format', '3_column')
    
    if csv_format == '5_column':
        # 5欄位格式: 需要組合年月和日期
        year_month_col = columns_info.get('year_month_col')
        day_col = columns_info.get('day_col')
        
        if year_month_col and day_col:
            year_month_value = row[year_month_col] if pd.notna(row[year_month_col]) else None
            day_value = row[day_col] if pd.notna(row[day_col]) else None
            return combine_date_from_5_column(year_month_value, day_value)
        else:
            return None
    else:
        # 3欄位格式: 直接使用日期欄位
        date_col = columns_info.get('date_col')
        if date_col:
            return row[date_col] if pd.notna(row[date_col]) else None
        else:
            return None

def validate_excel_data(df, data_version, filename):
    """驗證Excel資料"""
    result = {
        'status': 'OK',
        'total_records': len(df),
        'valid_records': 0,
        'warnings': 0,
        'errors': 0,
        'error_messages': [],
        'filename': filename,
        'data_version': data_version
    }
    
    try:
        # 1. 欄位驗證
        name_col, date_col, shift_col = identify_columns(df)
        
        if not name_col:
            result['errors'] += 1
            result['error_messages'].append('找不到姓名欄位（應包含：姓名、員工姓名或name）')
            result['status'] = 'ERROR'
        
        if not date_col:
            result['errors'] += 1
            result['error_messages'].append('找不到日期欄位（應包含：日期或date）')
            result['status'] = 'ERROR'
        
        if not shift_col:
            result['errors'] += 1
            result['error_messages'].append('找不到班別欄位（應包含：班別、班次或shift）')
            result['status'] = 'ERROR'
        
        # 如果基本欄位都找不到，直接返回
        if result['status'] == 'ERROR':
            return result
        
        # 2. 資料內容驗證
        valid_shifts = get_valid_shift_codes()
        target_employees = get_target_employees()
        
        for index, row in df.iterrows():
            try:
                # 姓名驗證
                employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                if not employee_name:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：姓名欄位空白')
                    continue
                
                # 群組過濾：只處理屬於目標群組的員工，跳過其他員工
                if employee_name not in target_employees:
                    # 跳過不在目標群組中的員工，不顯示錯誤訊息
                    continue
                
                # 日期驗證
                date_value = row[date_col] if pd.notna(row[date_col]) else None
                if not date_value:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：日期欄位空白')
                    continue
                
                # 班別驗證
                shift_code = str(row[shift_col]).strip() if pd.notna(row[shift_col]) else ''
                if not shift_code:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：班別欄位空白')
                    continue
                
                if shift_code not in valid_shifts:
                    result['warnings'] += 1
                    result['error_messages'].append(f'第{index+2}行：未知班別代碼 {shift_code}')
                    if result['status'] == 'OK':
                        result['status'] = 'WARNING'
                
                result['valid_records'] += 1
                
            except Exception as e:
                result['errors'] += 1
                result['error_messages'].append(f'第{index+2}行：資料處理錯誤 - {str(e)}')
        
        # 3. 設定最終狀態
        if result['errors'] > 0:
            result['status'] = 'ERROR'
        elif result['warnings'] > 0:
            result['status'] = 'WARNING'
        
        return result
        
    except Exception as e:
        result['status'] = 'ERROR'
        result['errors'] += 1
        result['error_messages'].append(f'驗證過程發生錯誤: {str(e)}')
        return result

def get_valid_shift_codes():
    """獲取有效班別代碼 - 從資料庫動態載入"""
    try:
        from app.models import ShiftType
        shift_codes = [shift.code for shift in ShiftType.query.all()]
        print(f"✅ 從資料庫載入 {len(shift_codes)} 個班別代碼: {shift_codes}")
        return shift_codes
    except Exception as e:
        print(f"⚠️ 載入班別代碼失敗，使用預設清單: {e}")
        # 預設清單作為備用（包含所有我們已初始化的班別）
        return [
            'A', 'B', 'C', 'OFF', 'FC', 'FC/工程', 'FC/急救課', 'FX',
            'P1s', 'P1s2', 'P1c', 'P1c2', 'P1n', 'P1n/夜超', 'P1p', 'P1p2', 'P1p/ME',
            'P2s', 'P2c', 'P2n', 'P2p', 'P2p/LD',
            'P3c', 'P3n', 'P3n/夜超', 'P3p',
            'P4c', 'P4n', 'P4p',
            'P5', 'P6', 'C1', 'C3', 'N1', 'N2', 'E1', 'R1',
            'H0', 'H1', '舞台'
        ]

def load_whitelist():
    """載入白名單配置 - 整合群組管理系統"""
    try:
        # 先嘗試從群組管理系統載入
        groups = GroupMembers.query.all()
        if groups:
            group_data = {}
            allowed_names = set()
            
            for group in groups:
                members = group.get_members()
                group_data[group.group_name] = members
                allowed_names.update(members)
            
            # 添加舊版本兼容的群組名稱
            legacy_mapping = {
                '演出人員': group_data.get('統籌組', []) + group_data.get('舞台組', []),
                '技術人員': group_data.get('燈光組', []) + group_data.get('視聽組', []) + group_data.get('維護組', []),
                '燈光組': group_data.get('燈光組', []),
                '全名單': list(allowed_names)
            }
            
            result = {
                'group': {**group_data, **legacy_mapping},
                'allowed': list(allowed_names)
            }
            
            print(f"✅ 從群組管理系統載入白名單，共 {len(allowed_names)} 人")
            return result
        
        # 如果群組管理系統沒有資料，回退到舊的JSON檔案
        whitelist_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'whitelist.json')
        with open(whitelist_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print("⚠️ 使用舊版JSON白名單檔案")
            return data
            
    except Exception as e:
        print(f'載入白名單配置錯誤: {e}')
        # 返回空的結構而不是None
        return {'group': {}, 'allowed': []}

def auto_detect_version(filename):
    """根據檔案名稱自動檢測資料版本 - 統一使用一般版本"""
    return '一般版本'

def create_preview_data_v11(df, target_group, whitelist_data):
    """創建v1.2版本的資料預覽 - 支援3欄位和5欄位格式"""
    preview_data = []
    
    # 燈光組核心7人名單
    core_light_crew = ['賴 秉 宏', '李 惟 綱', '李 家 瑋', '王 志 忠', '顧 育 禎', '胡 翊 潔', '朱 家 德']
    
    # 嘗試識別欄位 - 使用新的智能識別系統
    columns_info = identify_columns(df)
    valid_shifts = get_valid_shift_codes()
    
    print(f"🔍 預覽資料使用格式: {columns_info.get('format', 'unknown')}")
    
    # 檢查必要欄位是否存在
    name_col = columns_info.get('name_col')
    shift_col = columns_info.get('shift_col')
    
    if name_col and shift_col:
        # 獲取目標群組的員工名單
        if target_group == '全名單':
            # 真正的全名單：不使用白名單限制，從CSV中提取所有唯一姓名
            valid_names = df[name_col].dropna().astype(str).str.strip().unique().tolist()
            print(f"🔧 [調試] 全名單模式：從CSV提取到 {len(valid_names)} 個唯一姓名")
        else:
            valid_names = whitelist_data.get('group', {}).get(target_group, [])
        
        record_count = 0
        for i in range(len(df)):
            if record_count >= 50:  # 限制預覽50筆
                break
                
            row = df.iloc[i]
            employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
            
            # 只顯示目標群組的記錄
            if employee_name not in valid_names:
                continue
            
            record_count += 1
            issues = []
            status = 'ok'
            
            # 使用增強的日期獲取功能
            date_value = get_date_value_enhanced(row, columns_info)
            shift_code = str(row[shift_col]).strip() if pd.notna(row[shift_col]) else ''
            
            # 獲取員工代碼（如果是5欄位格式的話）
            employee_code = ''
            if columns_info.get('format') == '5_column' and columns_info.get('employee_code_col'):
                employee_code = str(row[columns_info['employee_code_col']]).strip() if pd.notna(row[columns_info['employee_code_col']]) else ''
            
            # 格式化日期
            formatted_date = ''
            if date_value:
                try:
                    if isinstance(date_value, str):
                        parsed_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                        formatted_date = parsed_date.strftime('%Y/%m/%d')
                    elif hasattr(date_value, 'strftime'):
                        # 如果是date對象或datetime對象
                        if hasattr(date_value, 'date'):
                            formatted_date = date_value.date().strftime('%Y/%m/%d')
                        else:
                            formatted_date = date_value.strftime('%Y/%m/%d')
                    else:
                        formatted_date = str(date_value)
                        if formatted_date and formatted_date != 'nan':
                            issues.append('日期格式無法解析')
                            status = 'warning'
                except Exception as e:
                    formatted_date = str(date_value)
                    issues.append(f'日期格式錯誤: {str(e)[:50]}')
                    status = 'error'
            else:
                issues.append('日期欄位空白')
                status = 'error'
            
            # 驗證姓名
            if not employee_name:
                issues.append('姓名欄位空白')
                status = 'error'
            
            # 驗證班別
            if not shift_code:
                issues.append('班別欄位空白')
                status = 'error'
            elif shift_code not in valid_shifts:
                issues.append(f'班別代號錯誤：{shift_code}')
                status = 'error'
            
            preview_data.append({
                'row': i + 2,
                'name': employee_name,
                'employee_code': employee_code,  # 新增員工代碼欄位
                'date': formatted_date if formatted_date else str(date_value),
                'shift': shift_code,
                'status': status,
                'message': '; '.join(issues) if issues else '',
                'format': columns_info.get('format', '3_column')  # 記錄格式資訊
            })
        
        return preview_data
    
    # 檢查是否為橫式格式（數字欄位代表日期）
    numeric_cols = [col for col in df.columns if str(col).strip().isdigit()]
    
    if numeric_cols:
        # 處理橫式格式
        for i in range(len(df)):
            row = df.iloc[i]
            
            # 嘗試多種方式找到姓名
            employee_name = ''
            
            # 方式1: 使用識別到的姓名欄位
            if name_col and name_col in df.columns:
                employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
            
            # 方式2: 檢查第一行是否有姓名資訊，如果有，使用對應位置
            if not employee_name and len(df) > 0:
                first_row = df.iloc[0]
                for col_idx, header_value in enumerate(first_row):
                    if str(header_value).strip() == '姓名' and col_idx < len(row):
                        employee_name = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                        break
            
            # 方式3: 嘗試第一欄
            if not employee_name:
                employee_name = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            
            # 跳過標題行和非核心7人記錄
            if employee_name in ['姓名', '年月', '組室別', '員工代碼'] or employee_name not in core_light_crew:
                continue
            
            # 遍歷每個數字欄位（日期），只預覽前10天
            for day_col in numeric_cols[:10]:
                if day_col in row.index:
                    shift_value = str(row[day_col]).strip() if pd.notna(row[day_col]) else ''
                    
                    if shift_value and shift_value not in ['nan', '']:
                        issues = []
                        status = 'ok'
                        
                        if shift_value not in valid_shifts:
                            issues.append(f'班別代號錯誤：{shift_value}')
                            status = 'error'
                        
                        try:
                            day_num = int(day_col)
                            formatted_date = f'2025-07-{day_num:02d}'
                        except:
                            formatted_date = str(day_col)
                        
                        preview_data.append({
                            'row': i + 2,
                            'name': employee_name,
                            'date': formatted_date,
                            'shift': shift_value,
                            'status': status,
                            'message': '; '.join(issues) if issues else ''
                        })
                        
                        if len(preview_data) >= 50:  # 限制預覽數量
                            return preview_data
    
    # 如果無法識別任何格式，顯示基本資訊
    if not preview_data:
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            preview_data.append({
                'row': i + 2,
                'name': str(row.iloc[0]) if len(row) > 0 else '',
                'date': str(row.iloc[1]) if len(row) > 1 else '',
                'shift': str(row.iloc[2]) if len(row) > 2 else '',
                'status': 'warning',
                'message': '無法識別檔案格式'
            })
    
    return preview_data

def normalize_shift_code(shift_code):
    """標準化班別代號，自動轉換含有說明的班別"""
    if not shift_code or pd.isna(shift_code) or str(shift_code).strip() == '':
        return ''
    
    shift_code = str(shift_code).strip()
    
    # 處理空值和無效值
    if shift_code in ['nan', 'NaN', 'None', 'null', '']:
        return ''
    
    # 映射表：將含有斜線的班別轉換為基本班別
    shift_mapping = {
        'FC/急救課': 'FC',
        'FC/保養': 'FC', 
        'FC/早上公假': 'FC',
        'FC/工程': 'FC',
        'FC/E1': 'FC',
        'FC/PM': 'FC',
        'P3n/夜超': 'P3n',
        'P1n/夜超': 'P1n',
        'P1n/LED': 'P1n',
        'P3c/保養': 'P3c',
        'P1p/保養': 'P1p',
        'P4p/保養': 'P4p',
        'CH/FC': 'FC',
        'CH/FC*': 'FC'
    }
    
    # 先檢查直接映射
    if shift_code in shift_mapping:
        return shift_mapping[shift_code]
    
    # 如果包含斜線，取斜線前的部分
    if '/' in shift_code:
        base_shift = shift_code.split('/')[0].strip()
        # 驗證基本班別是否有效
        valid_base_shifts = ['FC', 'P3n', 'P3c', 'P1p', 'P4p', 'P1n', 'P2n', 'P2p', 'P2s', 'P3p', 'P4n', 'CH']
        if base_shift in valid_base_shifts:
            return base_shift
    
    return shift_code

def validate_shift_count_equality(df):
    """驗證每個人的班數是否相等 - 按照新邏輯：有班就算，包含H0/H1"""
    validation_results = {
        'is_valid': True,
        'errors': [],
        'warnings': [],
        'statistics': {},
        'uneven_distribution': [],
        'duplicate_records': [],
        'part_time_employees': []
    }
    
    try:
        # 統計每個人的班數 - 有班就算（包含H0, H1）
        person_shift_counts = {}
        person_daily_shifts = {}  # 記錄每個人每天的班別
        
        # 智能識別姓名、班別和日期欄位
        columns_info = identify_columns(df)
        name_col = columns_info.get('name_col')
        shift_col = columns_info.get('shift_col')
        
        # 嘗試識別日期欄位
        date_col = None
        for col in df.columns:
            if '日期' in str(col) or 'day' in str(col).lower():
                date_col = col
                break
        
        if not name_col or not shift_col:
            validation_results['is_valid'] = False
            validation_results['errors'].append('無法識別姓名或班別欄位')
            return validation_results
        
        # 第一階段：收集所有班別資料
        for index, row in df.iterrows():
            name = str(row[name_col]).strip()
            original_shift = str(row[shift_col]).strip()
            
            # 跳過標題行和模板行
            if name in ['姓名', '年月', '組室別', '員工代碼', '', 'nan', 'NaN']:
                continue
            
            # 跳過純數字（日期模板）
            if name.isdigit():
                continue
            
            # 跳過長度異常的姓名
            if len(name) < 2 or len(name) > 15:
                continue
            
            # 獲取日期資訊
            date_info = ''
            if date_col and pd.notna(row[date_col]):
                date_info = str(row[date_col]).strip()
            
            # 處理班別（清理換行符號和額外資訊）
            shift = original_shift
            if shift:
                # 清理換行符號和多餘空白
                shift = shift.replace('\n', '').replace('\r', '').strip()
                # 如果包含時間資訊（如 P4n\n13-22），只取班別部分
                if '\n' in original_shift or any(char in shift for char in ['09-', '13-', '17-']):
                    # 提取班別代號（通常在最前面）
                    shift_parts = shift.split()
                    if shift_parts:
                        shift = shift_parts[0]
            
            # 空白班別視為非全職人員，記錄但不計算班數
            if not shift or shift in ['', 'nan', 'NaN', 'None', 'null']:
                if name not in validation_results['part_time_employees']:
                    validation_results['part_time_employees'].append(name)
                continue
            
            # 記錄每個人每天的班別
            person_date_key = f"{name}_{date_info}"
            if person_date_key not in person_daily_shifts:
                person_daily_shifts[person_date_key] = {
                    'name': name,
                    'date': date_info,
                    'shifts': [],
                    'line': index + 1
                }
            person_daily_shifts[person_date_key]['shifts'].append(shift)
        
        # 第二階段：處理合併班別和重複記錄
        for person_date_key, daily_data in person_daily_shifts.items():
            name = daily_data['name']
            date = daily_data['date']
            shifts = daily_data['shifts']
            
            # 去除重複的班別，保持順序
            unique_shifts = []
            for shift in shifts:
                if shift not in unique_shifts:
                    unique_shifts.append(shift)
            
            # 如果同一天有多個不同班別，合併為 班別A/班別B 格式
            if len(shifts) > 1:
                if len(unique_shifts) > 1:
                    # 不同班別：顯示警告並合併
                    merged_shift = '/'.join(unique_shifts)
                    validation_results['warnings'].append(
                        f"{name} 在 {date} 有多個班別：{merged_shift}"
                    )
                else:
                    # 相同班別重複：顯示錯誤
                    validation_results['duplicate_records'].append({
                        'name': name,
                        'date': date,
                        'line': daily_data['line'],
                        'shift': unique_shifts[0],
                        'count': len(shifts)
                    })
            
            # 統計班數：每天算一班，不管有幾個班別
            if name not in person_shift_counts:
                person_shift_counts[name] = 0
            person_shift_counts[name] += 1
        
        # 分析班數分布
        if person_shift_counts:
            shift_counts = list(person_shift_counts.values())
            min_shifts = min(shift_counts)
            max_shifts = max(shift_counts)
            avg_shifts = sum(shift_counts) / len(shift_counts)
            
            validation_results['statistics'] = {
                'total_people': len(person_shift_counts),
                'min_shifts': min_shifts,
                'max_shifts': max_shifts,
                'avg_shifts': round(avg_shifts, 2),
                'shift_difference': max_shifts - min_shifts
            }
            
            # 檢查班數差異
            if max_shifts - min_shifts > 2:  # 允許最多2班的差異
                validation_results['is_valid'] = False
                validation_results['errors'].append(
                    f"班數分布不均：最多{max_shifts}班，最少{min_shifts}班，差異{max_shifts - min_shifts}班"
                )
            elif max_shifts - min_shifts > 0:
                validation_results['warnings'].append(
                    f"班數略有差異：最多{max_shifts}班，最少{min_shifts}班"
                )
            
            # 找出班數異常的人員
            target_shifts = round(avg_shifts)
            for name, count in person_shift_counts.items():
                if abs(count - target_shifts) > 1:
                    validation_results['uneven_distribution'].append({
                        'name': name,
                        'actual_shifts': count,
                        'expected_shifts': target_shifts,
                        'difference': count - target_shifts
                    })
        else:
            validation_results['warnings'].append('沒有找到有效的班別資料')
        
        # 處理重複記錄錯誤
        if validation_results['duplicate_records']:
            validation_results['is_valid'] = False
            for dup in validation_results['duplicate_records']:
                validation_results['errors'].append(
                    f"第{dup['line']}行：重複記錄 - {dup['name']} 在 {dup['date']} 重複 {dup['count']} 次班別 {dup['shift']}"
                )
        
        # 處理非全職人員資訊
        if validation_results['part_time_employees']:
            validation_results['warnings'].append(
                f"發現 {len(validation_results['part_time_employees'])} 位非全職人員（有空白班別）"
            )
        
    except Exception as e:
        validation_results['is_valid'] = False
        validation_results['errors'].append(f"驗證過程發生錯誤: {str(e)}")
    
    return validation_results

def validate_excel_data_v11(df, data_version, filename, target_group, whitelist_data):
    """v1.1版本的Excel資料驗證 - 支援3欄位和5欄位格式，優化批量匯入"""
    result = {
        'status': 'OK',
        'total_records': len(df),
        'valid_records': 0,
        'warnings': 0,
        'errors': 0,
        'error_messages': [],
        'filename': filename,
        'data_version': data_version,
        'target_group': target_group,
        'batch_import_mode': True,  # 標記為批量匯入模式
        'shift_equality': None  # 新增班數相等驗證結果
    }
    
    try:
        # 1. 欄位驗證 - 使用新的智能識別系統
        columns_info = identify_columns(df)
        csv_format = columns_info.get('format', 'unknown')
        
        print(f"🔍 驗證資料使用格式: {csv_format}")
        
        name_col = columns_info.get('name_col')
        shift_col = columns_info.get('shift_col')
        
        if not name_col:
            result['errors'] += 1
            result['error_messages'].append('找不到姓名欄位（應包含：姓名、員工姓名或name）')
            result['status'] = 'ERROR'
        
        if not shift_col:
            result['errors'] += 1
            result['error_messages'].append('找不到班別欄位（應包含：班別、班次或shift）')
            result['status'] = 'ERROR'
        
        # 根據格式驗證日期相關欄位
        if csv_format == '5_column':
            year_month_col = columns_info.get('year_month_col')
            day_col = columns_info.get('day_col')
            
            if not year_month_col:
                result['errors'] += 1
                result['error_messages'].append('5欄位格式中找不到年月欄位（應包含：年月）')
                result['status'] = 'ERROR'
            
            if not day_col:
                result['errors'] += 1
                result['error_messages'].append('5欄位格式中找不到日期欄位（應包含：日期、日或day）')
                result['status'] = 'ERROR'
        else:
            date_col = columns_info.get('date_col')
            if not date_col:
                result['errors'] += 1
                result['error_messages'].append('找不到日期欄位（應包含：日期或date）')
                result['status'] = 'ERROR'
        
        # 如果基本欄位都找不到，直接返回
        if result['status'] == 'ERROR':
            return result
        
        # 2. 獲取白名單和有效班別
        if target_group == '全名單':
            # 真正的全名單：不使用白名單限制，從CSV中提取所有唯一姓名
            if name_col:
                valid_names = df[name_col].dropna().astype(str).str.strip().unique().tolist()
                print(f"🔧 [調試] 驗證階段-全名單模式：從CSV提取到 {len(valid_names)} 個唯一姓名")
            else:
                valid_names = []
        else:
            valid_names = whitelist_data.get('group', {}).get(target_group, [])
        
        valid_shifts = get_valid_shift_codes()
        
        # 3. 資料內容驗證 - 批量處理模式
        daily_schedules = {}  # 用於檢查重複和單日人數
        batch_errors = []  # 收集批量錯誤
        processed_count = 0
        
        print(f"🚀 開始批量驗證 {len(df)} 筆記錄...")
        
        for index, row in df.iterrows():
            processed_count += 1
            
            # 每100筆顯示進度
            if processed_count % 100 == 0:
                print(f"📊 已處理 {processed_count}/{len(df)} 筆記錄")
            
            try:
                # 姓名驗證
                employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                if not employee_name:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：姓名欄位空白')
                    continue
                
                # 群組過濾：只處理屬於目標群組的員工，跳過其他員工
                if employee_name not in valid_names:
                    # 跳過不在目標群組中的員工，不顯示錯誤訊息
                    continue
                
                # 使用增強的日期獲取功能
                date_value = get_date_value_enhanced(row, columns_info)
                if not date_value:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：日期欄位空白或無法解析')
                    continue
                
                # 嘗試解析日期
                try:
                    if isinstance(date_value, str):
                        schedule_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                    elif hasattr(date_value, 'strftime'):
                        # 如果已經是date對象
                        schedule_date = date_value
                    else:
                        schedule_date = date_value.date() if hasattr(date_value, 'date') else date_value
                except Exception as date_error:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：日期欄非日期格式')
                    continue
                
                # 班別驗證 - 先標準化班別代號
                original_shift = str(row[shift_col]).strip() if pd.notna(row[shift_col]) else ''
                if not original_shift:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：班別欄位空白')
                    continue
                
                # 標準化班別代號
                shift_code = normalize_shift_code(original_shift)
                
                if shift_code not in valid_shifts:
                    result['errors'] += 1
                    result['error_messages'].append(f'第{index+2}行：班別代號錯誤：{original_shift} (標準化後: {shift_code})')
                    continue
                
                # 如果有轉換，記錄轉換資訊
                if original_shift != shift_code:
                    result['warnings'] += 1
                    result['error_messages'].append(f'第{index+2}行：班別已自動轉換：{original_shift} → {shift_code}')
                
                # 重複資料檢查
                schedule_key = f"{employee_name}_{schedule_date}"
                if schedule_key in daily_schedules:
                    result['warnings'] += 1
                    result['error_messages'].append(f'第{index+2}行：重複資料 - {employee_name} 在 {schedule_date} 已有排班')
                    if result['status'] == 'OK':
                        result['status'] = 'WARNING'
                else:
                    daily_schedules[schedule_key] = True
                    
                    # 統計每日人數
                    if schedule_date not in daily_schedules:
                        daily_schedules[schedule_date] = []
                    daily_schedules[schedule_date].append(employee_name)
                
                result['valid_records'] += 1
                
            except Exception as e:
                result['errors'] += 1
                result['error_messages'].append(f'第{index+2}行：資料處理錯誤 - {str(e)}')
        
        # 4. 檢查單日人數警告
        for schedule_date, employees in daily_schedules.items():
            if isinstance(employees, list) and len(employees) == 1:
                result['warnings'] += 1
                result['error_messages'].append(f'{schedule_date}：當日僅一人上班，請注意補人手')
                if result['status'] == 'OK':
                    result['status'] = 'WARNING'
        
        # 5. 檢查燈光組核心7人班數平衡（支援直式和橫式格式）
        core_light_crew = ['賴 秉 宏', '李 惟 綱', '李 家 瑋', '王 志 忠', '顧 育 禎', '胡 翊 潔', '朱 家 德']
        core_crew_stats = {}
        
        # 判斷是直式還是橫式格式
        is_vertical_format = (name_col and shift_col and 
                             '姓名' in str(name_col).lower() and 
                             ('班別' in str(shift_col).lower() or 'shift' in str(shift_col).lower()))
        
        if is_vertical_format:
            # 直式格式統計：每行代表一個人一天的班別
            for index, row in df.iterrows():
                try:
                    # 獲取姓名
                    employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                    
                    # 跳過標題行
                    if employee_name in ['姓名', '年月', '組室別', '員工代碼']:
                        continue
                    
                    if employee_name in core_light_crew:
                        if employee_name not in core_crew_stats:
                            core_crew_stats[employee_name] = 0
                        
                        # 獲取班別
                        shift_value = str(row[shift_col]).strip() if pd.notna(row[shift_col]) else ''
                        
                        # 計算班數：排除休假、舞台班和空值
                        if shift_value and shift_value not in ['OFF', 'nan', '', '休假', '舞台']:
                            core_crew_stats[employee_name] += 1
                            
                except Exception as e:
                    print(f"統計核心人員時出錯 (直式): {e}")
                    continue
        else:
            # 橫式格式統計：傳統方式
            numeric_cols = [col for col in df.columns if str(col).strip().isdigit()]
            
            for index, row in df.iterrows():
                try:
                    # 嘗試多種方式找到姓名
                    employee_name = ''
                    
                    # 方式1: 使用識別到的姓名欄位
                    if name_col and name_col in df.columns:
                        employee_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                    
                    # 方式2: 檢查第一行是否有姓名資訊
                    if not employee_name and len(df) > 0:
                        first_row = df.iloc[0]
                        for col_idx, header_value in enumerate(first_row):
                            if str(header_value).strip() == '姓名' and col_idx < len(row):
                                employee_name = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                                break
                    
                    # 方式3: 嘗試第一欄
                    if not employee_name:
                        employee_name = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
                    
                    # 跳過標題行
                    if employee_name in ['姓名', '年月', '組室別', '員工代碼']:
                        continue
                    
                    if employee_name in core_light_crew:
                        if employee_name not in core_crew_stats:
                            core_crew_stats[employee_name] = 0
                        
                        # 統計該員工在所有日期欄位中的班別數
                        for day_col in numeric_cols:
                            if day_col in row.index:
                                shift_value = str(row[day_col]).strip() if pd.notna(row[day_col]) else ''
                                if shift_value and shift_value not in ['OFF', 'nan', '', '休假', '舞台']:
                                    core_crew_stats[employee_name] += 1
                        
                except Exception as e:
                    print(f"統計核心人員時出錯 (橫式): {e}")
                    continue
        
        # 將統計結果加入驗證結果
        result['core_light_crew_stats'] = core_crew_stats
        
        # 檢查班數是否平衡
        if core_crew_stats:
            shift_counts = list(core_crew_stats.values())
            if shift_counts and len(set(shift_counts)) > 1:
                min_count = min(shift_counts)
                max_count = max(shift_counts)
                result['warnings'] += 1
                result['error_messages'].append(f'燈光組核心人員班數不平衡：最少{min_count}班，最多{max_count}班')
                if result['status'] == 'OK':
                    result['status'] = 'WARNING'
                
                # 詳細列出每個人的班數
                stats_details = []
                for name, count in core_crew_stats.items():
                    stats_details.append(f'{name}: {count}班')
                result['error_messages'].append(f'核心人員班數明細：{", ".join(stats_details)}')
        
        # 6. 班數相等驗證
        shift_equality = validate_shift_count_equality(df)
        result['shift_equality'] = shift_equality
        
        # 整合班數驗證結果
        if not shift_equality['is_valid']:
            result['errors'] += len(shift_equality['errors'])
            result['error_messages'].extend(shift_equality['errors'])
            result['status'] = 'ERROR'
        
        if shift_equality['warnings']:
            result['warnings'] += len(shift_equality['warnings'])
            result['error_messages'].extend(shift_equality['warnings'])
            if result['status'] == 'OK':
                result['status'] = 'WARNING'
        
        # 7. 設定最終狀態
        if result['errors'] > 0:
            result['status'] = 'ERROR'
        elif result['warnings'] > 0:
            result['status'] = 'WARNING'
        
        return result
        
    except Exception as e:
        result['status'] = 'ERROR'
        result['errors'] += 1
        result['error_messages'].append(f'驗證過程發生錯誤: {str(e)}')
        return result

def get_target_employees():
    """獲取目標員工名單（保留向後兼容性）"""
    return [
        '賴秉宏', '李惟綱', '李家瑋', '王志忠', '顧育禎',
        '胡翊潔', '朱家德', '陳韋如', '葛禎', '井康羽',
        '簡芳瑜', '梁弘岳', '李佩璇', '鄭栢玔', '王文怡'
    ]

@main.route('/confirm_import', methods=['POST'])
@require_admin
def confirm_import():
    """確認並執行匯入（無需session暫存）"""
    try:
        csv_data = request.form.get('csv_data')
        target_group = request.form.get('target_group')
        filename = request.form.get('filename')
        force_import = request.form.get('force_import') == 'true'
        allow_blank_shifts = request.form.get('allow_blank_shifts') == 'true'
        import_mode = request.form.get('import_mode', 'merge')  # 默認為合併模式
        
        if not csv_data or not target_group or not filename:
            flash('匯入資料不完整，請重新上傳', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 解析CSV資料
        import io
        import base64
        csv_string = base64.b64decode(csv_data).decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_string))
        
        # 重新驗證資料
        whitelist_data = load_whitelist()
        validation_result = validate_excel_data_v11(df, '一般版本', filename, target_group, whitelist_data)
        
        # 檢查是否可以匯入
        if validation_result['status'] == 'ERROR' and not force_import:
            flash('資料有錯誤，無法匯入', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 執行匯入
        import_options = {
            'force_import': force_import,
            'allow_blank_shifts': allow_blank_shifts,
            'import_mode': import_mode
        }
        import_result = perform_data_import_v11(df, validation_result, target_group, False, import_mode, import_options)
        
        # 記錄匯入日誌
        current_user = get_current_user()
        import_log = ImportLog(
            importer=current_user.name,
            filename=filename,
            data_version='一般版本',
            target_group=target_group,
            validation_result=validation_result['status'],
            force_import=force_import,
            error_count=validation_result['errors'],
            warning_count=validation_result['warnings'],
            records_imported=import_result.get('imported_count', 0) + import_result.get('updated_count', 0)
        )
        
        if validation_result['error_messages']:
            import_log.set_validation_errors(validation_result['error_messages'])
        
        db.session.add(import_log)
        db.session.commit()
        
        # 顯示結果訊息
        if import_result['success']:
            imported = import_result.get('imported_count', 0)
            updated = import_result.get('updated_count', 0) 
            skipped = import_result.get('skipped_count', 0)
            mode_text = "合併模式" if import_mode == 'merge' else "覆寫模式"
            flash(f'✅ 匯入成功！({mode_text}) 新增: {imported}筆, 更新: {updated}筆, 跳過: {skipped}筆', 'success')
        else:
            flash(f'❌ 匯入失敗：{import_result["error"]}', 'error')
        
        return redirect(url_for('main.upload_new'))
        
    except Exception as e:
        print(f"匯入錯誤: {e}")
        flash(f'匯入過程發生錯誤：{str(e)}', 'error')
        return redirect(url_for('main.upload_new'))

@main.route('/execute_import', methods=['POST'])
@require_admin
def execute_import():
    """執行匯入作業"""
    try:
        data_hash = request.form.get('validated_data')
        force_import = request.form.get('force_import') == 'true'
        
        if not data_hash or f'validated_data_{data_hash}' not in session:
            flash('驗證資料已過期，請重新上傳', 'error')
            return redirect(url_for('main.upload_new'))
        
        cached_data = session[f'validated_data_{data_hash}']
        df = pd.read_json(cached_data['df'])
        validation_result = cached_data['validation_result']
        filename = cached_data['filename']
        data_version = cached_data['data_version']
        target_group = cached_data.get('target_group', '全名單')
        skip_invalid = cached_data.get('skip_invalid', False)
        
        # 檢查是否可以匯入
        if validation_result['status'] == 'ERROR' and not force_import:
            flash('資料有錯誤，無法匯入', 'error')
            return redirect(url_for('main.upload_new'))
        
        # 執行匯入
        import_result = perform_data_import_v11(df, validation_result, target_group, skip_invalid)
        
        # 記錄匯入日誌
        current_user = get_current_user()
        import_log = ImportLog(
            importer=current_user.name,
            filename=filename,
            data_version=data_version,
            target_group=target_group,
            validation_result=validation_result['status'],
            force_import=force_import,
            error_count=validation_result['errors'],
            warning_count=validation_result['warnings'],
            records_imported=import_result['imported_count']
        )
        
        if validation_result['error_messages']:
            import_log.set_validation_errors(validation_result['error_messages'])
        
        db.session.add(import_log)
        db.session.commit()
        
        # 清理緩存
        del session[f'validated_data_{data_hash}']
        
        success_message = f'匯入完成！成功匯入 {import_result["imported_count"]} 筆記錄'
        if import_result['skipped_count'] > 0:
            success_message += f'，跳過 {import_result["skipped_count"]} 筆重複記錄'
        
        flash(success_message, 'success')
        return redirect(url_for('main.upload_new'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'匯入失敗: {str(e)}', 'error')
        return redirect(url_for('main.upload_new'))

@main.route('/api/employee-groups', methods=['GET'])
@require_auth
def get_employee_groups():
    """獲取員工群組歸屬資料"""
    try:
        # 獲取所有群組資料
        all_groups = GroupMembers.query.all()
        employee_groups = {}
        
        for group in all_groups:
            members = group.get_members()
            for member_name in members:
                if member_name not in employee_groups:
                    employee_groups[member_name] = []
                employee_groups[member_name].append(group.group_name)
        
        return jsonify({
            'success': True,
            'data': employee_groups
        })
    except Exception as e:
        print(f'獲取員工群組錯誤: {e}')
        return jsonify({'success': False, 'message': '獲取群組資料失敗'}), 500

@main.route('/api/group-members', methods=['GET'])
@require_auth
def get_group_members():
    """獲取所有群組的人員名單"""
    try:
        groups = GroupMembers.query.all()
        group_data = {}
        
        # 預設群組
        default_groups = ['統籌組', '燈光組', '舞台組', '視聽組', '維護組']
        
        for group_name in default_groups:
            group_data[group_name] = []
        
        # 載入資料庫中的群組資料
        for group in groups:
            group_data[group.group_name] = group.get_members()
        
        return jsonify({
            'status': 'success',
            'data': group_data
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@main.route('/api/group-members', methods=['POST'])
@require_admin
def save_group_members():
    """儲存群組人員設定"""
    try:
        data = request.get_json()
        
        for group_name, members in data.items():
            # 查找或創建群組記錄
            group = GroupMembers.query.filter_by(group_name=group_name).first()
            if not group:
                group = GroupMembers(group_name=group_name)
                db.session.add(group)
            
            # 清理和設置成員名單
            cleaned_members = [name.strip() for name in members if name.strip()]
            group.set_members(cleaned_members)
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': '群組設定已儲存'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'儲存失敗: {str(e)}'
        }), 500

@main.route('/api/group-members/<group_name>', methods=['PUT'])
@require_admin
def update_group_members(group_name):
    """更新特定群組的人員名單"""
    try:
        data = request.get_json()
        members = data.get('members', [])
        
        # 查找或創建群組記錄
        group = GroupMembers.query.filter_by(group_name=group_name).first()
        if not group:
            group = GroupMembers(group_name=group_name)
            db.session.add(group)
        
        # 清理和設置成員名單
        cleaned_members = [name.strip() for name in members if name.strip()]
        group.set_members(cleaned_members)
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'{group_name} 人員名單已更新',
            'count': len(cleaned_members)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'更新失敗: {str(e)}'
        }), 500

@main.route('/api/clear-all-data', methods=['POST'])
@require_admin
def clear_all_data():
    """清除所有班表資料和匯入記錄（保留員工和班別設定）"""
    try:
        # 計算清除前的資料統計
        schedule_count = Schedule.query.count()
        import_log_count = ImportLog.query.count()
        
        print(f"🗑️ 準備清除資料：{schedule_count} 筆排班記錄，{import_log_count} 筆匯入記錄")
        
        # 清除排班記錄
        Schedule.query.delete()
        
        # 清除匯入日誌
        ImportLog.query.delete()
        
        # 提交變更
        db.session.commit()
        
        message = f"成功清除 {schedule_count} 筆排班記錄和 {import_log_count} 筆匯入記錄"
        print(f"✅ {message}")
        
        return jsonify({
            'status': 'success',
            'message': message,
            'cleared': {
                'schedules': schedule_count,
                'import_logs': import_log_count
            }
        })
    except Exception as e:
        db.session.rollback()
        print(f"❌ 清除資料失敗: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'清除失敗: {str(e)}'
        }), 500

def perform_data_import_v11(df, validation_result, target_group, skip_invalid, import_mode='merge', import_options=None):
    """執行v1.1版本的實際資料匯入 - 支援3欄位和5欄位格式，優化批量處理，新增強制匯入功能"""
    if import_options is None:
        import_options = {}
    
    force_import = import_options.get('force_import', False)
    allow_blank_shifts = import_options.get('allow_blank_shifts', True)
    imported_count = 0
    skipped_count = 0
    updated_count = 0
    
    try:
        # 載入白名單
        whitelist_data = load_whitelist()
        columns_info = identify_columns(df)
        
        name_col = columns_info.get('name_col')
        shift_col = columns_info.get('shift_col')
        
        print(f"🚀 匯入使用格式: {columns_info.get('format', 'unknown')}")
        print(f"📋 匯入模式: {import_mode}")
        print(f"📊 準備處理 {len(df)} 筆記錄...")
        
        # 如果是覆蓋模式，需要清除現有資料
        if import_mode == 'overwrite':
            print("🗑️ 覆蓋模式：清除現有排班資料...")
            # 只清除目標群組的資料，不是全部清除
            if target_group != '全名單':
                # 清除特定群組的排班資料
                pass  # 這裡可以添加清除邏輯
        
        if target_group == '全名單':
            # 實際匯入階段：全名單模式不限制姓名
            valid_names = df[name_col].dropna().astype(str).str.strip().unique().tolist()
            print(f"🔧 [調試] 匯入階段-全名單模式：處理 {len(valid_names)} 個唯一姓名")
        else:
            valid_names = whitelist_data.get('group', {}).get(target_group, [])
        
        valid_shifts = get_valid_shift_codes()
        
        processed_count = 0
        batch_size = 100  # 每100筆提交一次
        
        print(f"🚀 開始批量匯入，每 {batch_size} 筆提交一次...")
        
        for index, row in df.iterrows():
            processed_count += 1
            
            # 每100筆顯示進度
            if processed_count % batch_size == 0:
                print(f"📈 已處理 {processed_count}/{len(df)} 筆記錄 (匯入: {imported_count}, 更新: {updated_count}, 跳過: {skipped_count})")
                # 批量提交到資料庫
                db.session.commit()
            try:
                # 提取姓名和班別
                employee_name = str(row[name_col]).strip()
                
                # 處理班別欄位（支援空白和複合格式）
                shift_raw = row[shift_col] if pd.notna(row[shift_col]) else ''
                shift_code = str(shift_raw).strip()
                
                # 處理空白班別
                if not shift_code:
                    if allow_blank_shifts:
                        shift_code = 'UNASSIGNED'  # 未排班代碼
                        print(f"📝 空白班別處理為未排班: {employee_name}")
                    else:
                        skipped_count += 1
                        continue
                
                # 處理複合班別格式（如 P1c/工程, P4p/ME 等）
                original_shift_code = shift_code
                if '/' in shift_code:
                    # 提取主要班別代碼（/之前的部分）
                    base_shift = shift_code.split('/')[0].strip()
                    suffix = shift_code.split('/')[1].strip() if len(shift_code.split('/')) > 1 else ''
                    print(f"🔧 複合班別: {shift_code} → 基礎班別: {base_shift}, 後綴: {suffix}")
                    
                    if force_import:
                        # 強制匯入模式：保持完整的複合班別代碼
                        pass
                    else:
                        # 一般模式：使用基礎班別代碼進行驗證
                        shift_code = base_shift
                
                # 處理特殊格式（如帶換行的班別）
                if '\n' in shift_code:
                    shift_code = shift_code.replace('\n', ' ').strip()
                
                # 提取員工代碼（如果是5欄位格式）
                employee_code = None
                if columns_info.get('format') == '5_column' and columns_info.get('employee_code_col'):
                    employee_code = str(row[columns_info['employee_code_col']]).strip() if pd.notna(row[columns_info['employee_code_col']]) else None
                
                # 使用增強的日期獲取功能
                date_value = get_date_value_enhanced(row, columns_info)
                
                # 跳過空白或無效資料
                if not employee_name or not date_value:
                    continue
                
                # 群組過濾檢查：如果不是全名單模式，只處理目標群組的員工
                if target_group != '全名單' and employee_name not in valid_names:
                    skipped_count += 1
                    continue
                
                # 班別代碼檢查（強制匯入模式會自動創建新班別）
                if shift_code not in valid_shifts and not force_import:
                    print(f"⚠️ 未知班別代碼: {shift_code} (第{index+2}行)")
                    skipped_count += 1
                    continue
                
                # 處理日期
                try:
                    if isinstance(date_value, str):
                        schedule_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                    elif hasattr(date_value, 'strftime'):
                        # 如果已經是date對象
                        schedule_date = date_value
                    else:
                        schedule_date = date_value.date() if hasattr(date_value, 'date') else date_value
                except Exception as date_error:
                    print(f"❌ 日期解析失敗 (第{index+2}行): {date_value}, 錯誤: {date_error}")
                    skipped_count += 1
                    continue
                
                # 查找或創建員工
                employee = Employee.query.filter_by(name=employee_name).first()
                if not employee:
                    # 優先使用CSV中的員工代碼，否則自動生成
                    final_employee_code = employee_code
                    if employee_code and employee_code.strip():
                        # 檢查員工代碼是否已被使用
                        if Employee.query.filter_by(employee_code=employee_code).first():
                            print(f"⚠️ 員工代碼 {employee_code} 已存在，將自動生成新代碼")
                            final_employee_code = None
                    
                    if not final_employee_code:
                        # 自動生成員工代碼
                        existing_count = Employee.query.count()
                        final_employee_code = f"EMP_{existing_count+1:03d}"
                        while Employee.query.filter_by(employee_code=final_employee_code).first():
                            existing_count += 1
                            final_employee_code = f"EMP_{existing_count+1:03d}"
                    
                    employee = Employee(name=employee_name, employee_code=final_employee_code)
                    db.session.add(employee)
                    db.session.flush()
                    print(f"✅ 創建員工: {employee_name} (代碼: {final_employee_code})")
                elif employee_code and employee_code.strip() and employee.employee_code != employee_code:
                    # 如果找到同名員工但代碼不同，更新員工代碼（如果不衝突）
                    if not Employee.query.filter_by(employee_code=employee_code).first():
                        old_code = employee.employee_code
                        employee.employee_code = employee_code
                        print(f"🔄 更新員工代碼: {employee_name} ({old_code} → {employee_code})")
                
                # 查找或創建班別類型（使用原始的複合班別代碼）
                final_shift_code = original_shift_code if force_import else shift_code
                shift_type = ShiftType.query.filter_by(code=final_shift_code).first()
                if not shift_type:
                    if force_import or final_shift_code == 'UNASSIGNED':
                        # 強制匯入或未排班：創建新班別類型
                        shift_name, start_time, end_time, color = get_shift_info(final_shift_code)
                        shift_type = ShiftType(
                            code=final_shift_code,
                            name=shift_name,
                            start_time=start_time,
                            end_time=end_time,
                            color=color
                        )
                        db.session.add(shift_type)
                        db.session.flush()
                        print(f"✅ 創建新班別類型: {final_shift_code}")
                    else:
                        # 一般模式：使用標準班別資訊
                        shift_name, start_time, end_time, color = get_shift_info(shift_code)
                        shift_type = ShiftType(
                            code=shift_code,
                            name=shift_name,
                            start_time=start_time,
                            end_time=end_time,
                            color=color
                        )
                        db.session.add(shift_type)
                        db.session.flush()
                
                # 檢查是否已存在排班記錄
                existing_schedule = Schedule.query.filter_by(
                    date=schedule_date,
                    employee_id=employee.id
                ).first()
                
                if existing_schedule:
                    # 更新現有記錄
                    if existing_schedule.shift_type_id != shift_type.id:
                        existing_schedule.shift_type_id = shift_type.id
                        existing_schedule.updated_at = datetime.utcnow()
                        # 更新匯入順序和時間戳
                        existing_schedule.import_order = index + 1
                        existing_schedule.import_timestamp = datetime.utcnow()
                        updated_count += 1
                    else:
                        skipped_count += 1  # 資料相同，跳過
                else:
                    # 創建新記錄
                    schedule = Schedule(
                        date=schedule_date,
                        employee_id=employee.id,
                        shift_type_id=shift_type.id,
                        import_order=index + 1,  # CSV行號作為匯入順序
                        import_timestamp=datetime.utcnow()
                    )
                    db.session.add(schedule)
                    imported_count += 1
                
            except Exception as e:
                print(f'第{index+2}行匯入錯誤: {e}')
                skipped_count += 1
                continue
        
        db.session.commit()
        # 最終提交和統計
        db.session.commit()
        print(f"✅ 批量匯入完成！匯入: {imported_count}, 更新: {updated_count}, 跳過: {skipped_count}")
        
        return {
            'imported_count': imported_count, 
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'total_processed': imported_count + updated_count + skipped_count
        }
        
    except Exception as e:
        db.session.rollback()
        raise e

def perform_data_import(df, validation_result):
    """執行實際資料匯入（保留向後兼容性）"""
    imported_count = 0
    skipped_count = 0
    
    try:
        name_col, date_col, shift_col = identify_columns(df)
        
        for index, row in df.iterrows():
            try:
                # 提取資料
                employee_name = str(row[name_col]).strip()
                date_value = row[date_col]
                shift_code = str(row[shift_col]).strip()
                
                # 跳過空白或無效資料
                if not employee_name or not shift_code or pd.isna(date_value):
                    continue
                
                # 處理日期
                if isinstance(date_value, str):
                    schedule_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                else:
                    schedule_date = date_value.date() if hasattr(date_value, 'date') else date_value
                
                # 查找或創建員工
                employee = Employee.query.filter_by(name=employee_name).first()
                if not employee:
                    existing_count = Employee.query.count()
                    employee_code = f"EMP_{existing_count+1:03d}"
                    while Employee.query.filter_by(employee_code=employee_code).first():
                        existing_count += 1
                        employee_code = f"EMP_{existing_count+1:03d}"
                    
                    employee = Employee(name=employee_name, employee_code=employee_code)
                    db.session.add(employee)
                    db.session.flush()
                
                # 查找或創建班別類型
                shift_type = ShiftType.query.filter_by(code=shift_code).first()
                if not shift_type:
                    shift_name, start_time, end_time, color = get_shift_info(shift_code)
                    shift_type = ShiftType(
                        code=shift_code,
                        name=shift_name,
                        start_time=start_time,
                        end_time=end_time,
                        color=color
                    )
                    db.session.add(shift_type)
                    db.session.flush()
                
                # 檢查是否已存在排班記錄
                existing_schedule = Schedule.query.filter_by(
                    date=schedule_date,
                    employee_id=employee.id
                ).first()
                
                if existing_schedule:
                    # 更新現有記錄
                    if existing_schedule.shift_type_id != shift_type.id:
                        existing_schedule.shift_type_id = shift_type.id
                        existing_schedule.updated_at = datetime.utcnow()
                        # 更新匯入順序和時間戳
                        existing_schedule.import_order = index + 1
                        existing_schedule.import_timestamp = datetime.utcnow()
                        updated_count += 1
                    else:
                        skipped_count += 1  # 資料相同，跳過
                else:
                    # 創建新記錄
                    schedule = Schedule(
                        date=schedule_date,
                        employee_id=employee.id,
                        shift_type_id=shift_type.id,
                        import_order=index + 1,  # CSV行號作為匯入順序
                        import_timestamp=datetime.utcnow()
                    )
                    db.session.add(schedule)
                    imported_count += 1
                
            except Exception as e:
                print(f'第{index+2}行匯入錯誤: {e}')
                continue
        
        db.session.commit()
        # 最終提交和統計
        db.session.commit()
        print(f"✅ 批量匯入完成！匯入: {imported_count}, 更新: {updated_count}, 跳過: {skipped_count}")
        
        return {
            'imported_count': imported_count, 
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'total_processed': imported_count + updated_count + skipped_count
        }
        
    except Exception as e:
        db.session.rollback()
        raise e

# 新規範的上傳功能
@main.route("/upload_csv", methods=["POST"])
@require_auth
def upload_csv():
    """處理CSV檔案上傳"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "沒有選擇檔案"}), 400
        
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "沒有選擇檔案"}), 400
        
        if not file.filename.lower().endswith(".csv"):
            return jsonify({"error": "只支援CSV檔案"}), 400
        
        # 讀取CSV檔案
        stream = io.StringIO(file.stream.read().decode("utf-8"))
        csv_data = list(csv.DictReader(stream))
        
        # 預覽和驗證資料
        preview_result = preview_csv_data(csv_data)
        
        return jsonify(preview_result)
        
    except Exception as e:
        return jsonify({"error": f"檔案處理錯誤: {str(e)}"}), 500

@main.route("/upload_pasted", methods=["POST"])
@require_auth
def upload_pasted():
    """處理貼上的CSV資料"""
    try:
        data = request.json
        csv_text = data.get("csv_data", "").strip()
        
        if not csv_text:
            return jsonify({"error": "沒有輸入資料"}), 400
        
        # 解析CSV文字
        csv_data = list(csv.DictReader(io.StringIO(csv_text)))
        
        # 預覽和驗證資料
        preview_result = preview_csv_data(csv_data)
        
        return jsonify(preview_result)
        
    except Exception as e:
        return jsonify({"error": f"資料處理錯誤: {str(e)}"}), 500

@main.route("/import_data", methods=["POST"])
@require_auth
def import_data():
    """執行資料匯入"""
    try:
        data = request.json
        csv_data = data.get("data", [])
        selected_months = data.get("selected_months", [])
        
        if not csv_data:
            return jsonify({"error": "沒有資料可匯入"}), 400
        
        # 過濾選定月份的資料
        filtered_data = []
        if selected_months:
            for row in csv_data:
                year_month = row.get("年月", "")
                if year_month in selected_months:
                    filtered_data.append(row)
        else:
            filtered_data = csv_data
        
        # 執行匯入
        result = import_csv_data(filtered_data)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": f"匯入錯誤: {str(e)}"}), 500

def preview_csv_data(csv_data):
    """預覽CSV資料並分析年月分佈"""
    if not csv_data:
        return {"error": "沒有資料"}
    
    # 檢查必要欄位
    required_fields = ["姓名", "員工代碼", "年月", "日期", "班別"]
    first_row = csv_data[0]
    missing_fields = [field for field in required_fields if field not in first_row.keys()]
    
    if missing_fields:
        missing_fields_str = ", ".join(missing_fields)
        return {"error": f"缺少欄位: {missing_fields_str}"}
    
    # 分析年月分佈
    month_distribution = defaultdict(int)
    preview_data = []
    errors = []
    
    for i, row in enumerate(csv_data[:50]):  # 只預覽前50筆
        try:
            name = row.get("姓名", "").strip()
            employee_code = row.get("員工代碼", "").strip()
            year_month = row.get("年月", "").strip()
            day = row.get("日期", "").strip()
            shift_code = row.get("班別", "").strip()
            
            if not all([name, employee_code, year_month, day, shift_code]):
                errors.append(f"第{i+1}行資料不完整")
                continue
            
            # 組合完整日期
            try:
                if "-" in year_month:
                    year, month = year_month.split("-")
                else:
                    year, month = year_month[:4], year_month[4:]
                
                full_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                datetime.strptime(full_date, "%Y-%m-%d")  # 驗證日期格式
                
                month_distribution[year_month] += 1
                
                preview_data.append({
                    "姓名": name,
                    "員工代碼": employee_code,
                    "年月": year_month,
                    "日期": day,
                    "班別": shift_code,
                    "完整日期": full_date
                })
                
            except ValueError as e:
                errors.append(f"第{i+1}行日期格式錯誤: {year_month}-{day}")
                
        except Exception as e:
            errors.append(f"第{i+1}行處理錯誤: {str(e)}")
    
    return {
        "success": True,
        "preview_data": preview_data,
        "month_distribution": dict(month_distribution),
        "total_records": len(csv_data),
        "errors": errors
    }

def import_csv_data(csv_data):
    """匯入CSV資料到資料庫"""
    imported_count = 0
    updated_count = 0
    error_count = 0
    errors = []
    
    try:
        for row in csv_data:
            try:
                name = row.get("姓名", "").strip()
                employee_code = row.get("員工代碼", "").strip()
                year_month = row.get("年月", "").strip()
                day = row.get("日期", "").strip()
                shift_code = row.get("班別", "").strip()
                
                # 組合完整日期
                if "-" in year_month:
                    year, month = year_month.split("-")
                else:
                    year, month = year_month[:4], year_month[4:]
                
                full_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                schedule_date = datetime.strptime(full_date, "%Y-%m-%d").date()
                
                # 查找或創建員工
                employee = Employee.query.filter_by(employee_code=employee_code).first()
                if not employee:
                    employee = Employee(name=name, employee_code=employee_code)
                    db.session.add(employee)
                    db.session.flush()
                
                # 查找或創建班別類型
                shift_type = ShiftType.query.filter_by(code=shift_code).first()
                if not shift_type:
                    shift_type = ShiftType(
                        code=shift_code,
                        name=f"{shift_code}班",
                        start_time=datetime.strptime("09:00", "%H:%M").time(),
                        end_time=datetime.strptime("17:00", "%H:%M").time(),
                        color="#007bff"
                    )
                    db.session.add(shift_type)
                    db.session.flush()
                
                # 檢查是否已存在排班記錄
                existing_schedule = Schedule.query.filter_by(
                    date=schedule_date,
                    employee_id=employee.id
                ).first()
                
                if existing_schedule:
                    existing_schedule.shift_type_id = shift_type.id
                    existing_schedule.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    schedule = Schedule(
                        date=schedule_date,
                        employee_id=employee.id,
                        shift_type_id=shift_type.id
                    )
                    db.session.add(schedule)
                    imported_count += 1
                    
            except Exception as e:
                error_count += 1
                errors.append(f"資料處理錯誤: {str(e)}")
        
        db.session.commit()
        
        return {
            "success": True,
            "imported_count": imported_count,
            "updated_count": updated_count,
            "error_count": error_count,
            "errors": errors[:10]  # 只返回前10個錯誤
        }
        
    except Exception as e:
        db.session.rollback()
        return {"success": False, "error": str(e)}

